"""
Flask PI Management System
Foreign Trade Proforma Invoice Generator
Run with: python app.py
"""

import os
import sys
import uuid
import zipfile
import shutil
from datetime import datetime, date, timedelta
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, send_file, jsonify, current_app, session, make_response
)
from sqlalchemy import func
from sqlalchemy.orm import joinedload
from models import db, Customer, Product, PI, PIItem, Salesperson, User, Account, Payment, Expense, Supplier, Procurement
from pdf_generator import generate_pi_pdf
from excel_generator import generate_pi_excel
from supplier_statement_pdf import generate_supplier_statement

# ── Configuration ────────────────────────────────────────────────────
# Detect the app root directory (where this file lives)
APP_ROOT = os.path.dirname(os.path.abspath(__file__))

import blob_sync

def _migrate_db():
    """Auto-add missing columns to existing tables without data loss."""
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    expected = {
        'customers': {'salesperson': 'VARCHAR(100)', 'created_at': 'DATETIME', 'total_deal_usd': 'FLOAT', 'image': 'VARCHAR(500)'},
        'products': {'image': 'VARCHAR(500)', 'chinese_name': 'VARCHAR(200)', 'unit_price_rmb': 'FLOAT'},
        'pis': {'salesperson': 'VARCHAR(100)', 'currency': 'VARCHAR(3)', 'company': 'VARCHAR(50)', 'excel_path': 'VARCHAR(500)', 'paid': 'BOOLEAN', 'received_amount': 'FLOAT', 'shipping_address': 'TEXT', 'actual_shipping_cost': 'FLOAT', 'procurement_confirmed': 'BOOLEAN'},
        'salespersons': {'phone': 'VARCHAR(50)', 'email': 'VARCHAR(200)', 'dingtalk_user_id': 'VARCHAR(100)'},
        'payments': {'order_no': 'VARCHAR(200)'},
        'accounts': {},  # table auto-created by create_all
        'suppliers': {},  # table auto-created by create_all
        'procurements': {},  # table auto-created by create_all
        'users': {'active': ('BOOLEAN', '1')},
    }
    for table, columns in expected.items():
        if not inspector.has_table(table): continue
        existing_cols = {c['name'] for c in inspector.get_columns(table)}
        for col_name, col_info in columns.items():
            if col_name not in existing_cols:
                # col_info can be a string (type only, DEFAULT '') or a tuple (type, default_sql)
                if isinstance(col_info, tuple):
                    col_type, col_default = col_info
                else:
                    col_type, col_default = col_info, "''"
                try:
                    db.session.execute(text(f"ALTER TABLE {table} ADD COLUMN {col_name} {col_type} DEFAULT {col_default}"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()


# Company info — edit these to match your business
COMPANY_CONFIG = {
    'name': 'CHANGZHOU KLISTA INTERNATIONAL TRADE CO., LTD.',
    'address': 'No. 158 Jinchuang Road, Yaoguan Town, Changzhou City, China',
    'phone': '+86 17712333882',
    'email': 'fantastic10086@gmail.com',
    'dingtalk_webhook': '',  # DingTalk robot webhook URL
}


SETTINGS_FILE = os.path.join(APP_ROOT, 'settings.json')

def _load_settings():
    import json as _json
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r') as f:
                return _json.load(f)
        except Exception:
            pass
    return {}

def _save_settings(data):
    import json as _json
    with open(SETTINGS_FILE, 'w') as f:
        _json.dump(data, f, indent=2)

def _get_webhook():
    """Get DingTalk webhook URL from settings file or env."""
    settings = _load_settings()
    return settings.get('dingtalk_webhook', COMPANY_CONFIG.get('dingtalk_webhook', '')).strip()


def _send_dingtalk(title, text):
    """Send a markdown message to DingTalk via webhook."""
    import json as _json
    from urllib import request as _req
    webhook = _get_webhook()
    if not webhook:
        return False
    payload = _json.dumps({
        'msgtype': 'markdown',
        'markdown': {'title': title, 'text': text},
    }).encode('utf-8')
    try:
        # Use unverified SSL context for compatibility (NAS/macOS Python 3.7)
        import ssl as _ssl
        ctx = _ssl._create_unverified_context()
        _req.urlopen(_req.Request(webhook, data=payload, headers={'Content-Type': 'application/json'}), timeout=10, context=ctx)
        return True
    except Exception:
        return False


_DT_CACHE = {'token': '', 'expires': 0, 'unionid': {}}

def _dingtalk_token():
    """Get DingTalk access_token using AppKey/AppSecret."""
    import json as _json, time as _time
    from urllib import request as _req
    settings = _load_settings()
    appkey = settings.get('dingtalk_appkey', '')
    appsecret = settings.get('dingtalk_appsecret', '')
    if not appkey or not appsecret:
        return ''
    now = _time.time()
    if _DT_CACHE['token'] and _DT_CACHE['expires'] > now:
        return _DT_CACHE['token']
    try:
        import ssl as _ssl
        ctx = _ssl._create_unverified_context()
        url = f'https://oapi.dingtalk.com/gettoken?appkey={appkey}&appsecret={appsecret}'
        resp = _req.urlopen(_req.Request(url), timeout=10, context=ctx)
        data = _json.loads(resp.read().decode('utf-8'))
        if data.get('errcode') == 0:
            _DT_CACHE['token'] = data['access_token']
            _DT_CACHE['expires'] = now + data.get('expires_in', 7200) - 300
            return _DT_CACHE['token']
    except Exception:
        pass
    return ''


def _dingtalk_userid_to_unionid(user_id):
    """Convert DingTalk userid to unionid using the user/get API."""
    import json as _json
    from urllib import request as _req
    import ssl as _ssl

    # Check cache
    cached = _DT_CACHE['unionid'].get(user_id)
    if cached:
        return cached

    token = _dingtalk_token()
    if not token or not user_id:
        return ''

    try:
        ctx = _ssl._create_unverified_context()
        url = f'https://oapi.dingtalk.com/user/get?access_token={token}&userid={user_id}'
        resp = _req.urlopen(_req.Request(url), timeout=10, context=ctx)
        data = _json.loads(resp.read().decode('utf-8'))
        if data.get('errcode') == 0:
            unionid = data.get('unionid', '')
            if unionid:
                _DT_CACHE['unionid'][user_id] = unionid
                return unionid
    except Exception:
        pass
    return ''


def _dingtalk_create_task(executor_user_ids, subject, description):
    """Create a DingTalk task (待办).

    executor_user_ids: list of DingTalk userids. First one used as creator.
    The function auto-converts userid → unionid internally.
    """
    import json as _json, time as _time
    from urllib import request as _req

    if not executor_user_ids:
        return ''

    # Convert all userids to unionids (dedup, keep order)
    union_ids = []
    seen = set()
    for uid in executor_user_ids:
        if not uid:
            continue
        u = _dingtalk_userid_to_unionid(uid)
        if u and u not in seen:
            union_ids.append(u)
            seen.add(u)
    if not union_ids:
        return ''

    token = _dingtalk_token()
    if not token:
        return ''

    payload = {
        'sourceId': f'pi_manager_{int(_time.time() * 1000)}',
        'subject': subject[:200],
        'description': description[:4000],
        'creatorId': union_ids[0],
        'executorIds': union_ids,
    }

    try:
        import ssl as _ssl
        ctx = _ssl._create_unverified_context()
        url = f'https://api.dingtalk.com/v1.0/todo/users/{union_ids[0]}/tasks?operatorId={union_ids[0]}'
        data = _json.dumps(payload).encode('utf-8')
        req = _req.Request(url, data=data, method='POST')
        req.add_header('Content-Type', 'application/json')
        req.add_header('x-acs-dingtalk-access-token', token)
        resp = _req.urlopen(req, timeout=15, context=ctx)
        result = _json.loads(resp.read().decode('utf-8'))
        if result.get('id'):
            return result['id']
    except Exception:
        pass
    return ''


def _dingtalk_send_notification(user_ids, title, text):
    """Send a work notification (工作通知) to DingTalk users via the enterprise app.

    user_ids: single user_id string or list of user_id strings.
    """
    import json as _json
    from urllib import request as _req
    import ssl as _ssl

    settings = _load_settings()
    agent_id = settings.get('dingtalk_agent_id', '').strip()
    if not agent_id or not user_ids:
        return False

    # Support both single user_id and list of user_ids
    if isinstance(user_ids, list):
        user_id_str = ','.join([u for u in user_ids if u])
    else:
        user_id_str = user_ids
    if not user_id_str:
        return False

    token = _dingtalk_token()
    if not token:
        return False

    payload = {
        'agent_id': agent_id,
        'userid_list': user_id_str,
        'msg': {
            'msgtype': 'markdown',
            'markdown': {
                'title': title[:50],
                'text': text[:4000],
            }
        }
    }

    try:
        ctx = _ssl._create_unverified_context()
        url = 'https://oapi.dingtalk.com/topapi/message/corpconversation/asyncsend_v2?access_token=' + token
        data = _json.dumps(payload).encode('utf-8')
        req = _req.Request(url, data=data, method='POST')
        req.add_header('Content-Type', 'application/json')
        resp = _req.urlopen(req, timeout=15, context=ctx)
        result = _json.loads(resp.read().decode('utf-8'))
        return result.get('errcode') == 0
    except Exception:
        pass
    return False


def _dingtalk_send_file(user_ids, filepath):
    """Upload and send a file (PDF) to DingTalk users via work notification.

    user_ids: single user_id string or list of user_id strings.
    filepath: absolute path to the file to send.
    """
    import json as _json
    from urllib import request as _req
    import ssl as _ssl

    if not user_ids or not filepath or not os.path.exists(filepath):
        return False

    settings = _load_settings()
    agent_id = settings.get('dingtalk_agent_id', '').strip()
    if not agent_id:
        return False

    token = _dingtalk_token()
    if not token:
        return False

    # Upload the file to DingTalk media
    media_id = _dingtalk_upload_media(token, filepath)
    if not media_id:
        return False

    # Re-fetch token (upload may have refreshed it)
    token = _dingtalk_token()
    if not token:
        return False

    # Support both single user_id and list
    if isinstance(user_ids, list):
        user_id_str = ','.join([u for u in user_ids if u])
    else:
        user_id_str = user_ids
    if not user_id_str:
        return False

    try:
        ctx = _ssl._create_unverified_context()
        url = 'https://oapi.dingtalk.com/topapi/message/corpconversation/asyncsend_v2?access_token=' + token
        payload = {
            'agent_id': agent_id,
            'userid_list': user_id_str,
            'msg': {
                'msgtype': 'file',
                'file': {'media_id': media_id}
            }
        }
        data = _json.dumps(payload).encode('utf-8')
        req = _req.Request(url, data=data, method='POST')
        req.add_header('Content-Type', 'application/json')
        resp = _req.urlopen(req, timeout=15, context=ctx)
        result = _json.loads(resp.read().decode('utf-8'))
        return result.get('errcode') == 0
    except Exception:
        pass
    return False


def _dingtalk_upload_media(token, filepath):
    """Upload a file to DingTalk, return media_id."""
    import json as _json
    from urllib import request as _req

    import ssl as _ssl
    ctx = _ssl._create_unverified_context()

    # Read file directly since we know it's a PDF
    try:
        with open(filepath, 'rb') as f:
            file_data = f.read()
    except Exception:
        return ''

    boundary = '----FormBoundary7MA4YWxk'
    body = []
    body.append('--' + boundary)
    body.append('Content-Disposition: form-data; name="media"; filename="pi.pdf"')
    body.append('Content-Type: application/pdf')
    body.append('')
    # Add bytes
    body_bytes = []
    for line in body:
        body_bytes.append(line.encode('utf-8'))
    body_bytes.append(file_data)
    body_bytes.append(('--' + boundary + '--').encode('utf-8'))
    full_body = b'\r\n'.join(body_bytes)

    url = f'https://oapi.dingtalk.com/media/upload?access_token={token}&type=file'
    req = _req.Request(url, data=full_body)
    req.add_header('Content-Type', 'multipart/form-data; boundary=' + boundary)
    try:
        resp = _req.urlopen(req, timeout=30, context=ctx)
        result = _json.loads(resp.read().decode('utf-8'))
        if result.get('errcode') == 0:
            return result.get('media_id', '')
    except Exception:
        pass
    return ''


def _seed_products_from_csv():
    """Seed products from bundled CSV files on first startup."""
    import csv, re
    for fname in ['产品信息202606111_1.csv', '产品信息202606111_2.csv', '产品信息202606111_3.csv']:
        csv_path = os.path.join(APP_ROOT, fname)
        if not os.path.exists(csv_path):
            continue
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader)
            for row in reader:
                try:
                    name = (row[0] or '').strip().replace('\n', ' ').replace('\r', '')[:200]
                    if not name: continue
                    code = (row[1] or '').strip()[:100]
                    spec = (row[2] or '').strip()[:200]
                    price_str = (row[5] or '').strip()
                    price = 0.0
                    if price_str:
                        nums = re.findall(r'[\d.]+', price_str.replace(',', ''))
                        if nums:
                            try: price = float(nums[0])
                            except: pass
                    db.session.add(Product(name=name, product_code=code, specification=spec, unit_price=price))
                except Exception:
                    pass
    db.session.commit()


def _seed_customers_from_csv():
    """Seed customers from bundled CSV file on first startup."""
    import csv
    csv_path = os.path.join(APP_ROOT, 'CUSTOMER_EXPORT_56677547_1_1781157273.csv')
    if not os.path.exists(csv_path):
        return
    count = 0
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            try:
                name = (row[0] or '').strip()
                if not name or name.startswith('NA_') or name.startswith('nocompany_'):
                    contact = (row[3] or '').strip()
                    name = contact if (contact and not contact.startswith('NA_')) else f'Customer_{row[2]}'
                contact_person = (row[3] or '').strip()[:100]
                email = (row[4] or '').strip() if len(row) > 4 else ''
                salesperson = (row[5] or '').strip() if len(row) > 5 else ''
                country = (row[6] or '').strip() if len(row) > 6 else ''
                address = (row[7] or '').strip() if len(row) > 7 else ''
                phone_raw = (row[9] or '').strip() if len(row) > 9 else ''
                phone = ''
                if phone_raw:
                    for p in phone_raw.replace("'", '').split(';'):
                        p = p.strip().strip('"+').strip()
                        if p and any(c.isdigit() for c in p):
                            phone = p[:50]; break
                deal_str = row[-1].strip() if row[-1] else '0'
                try:
                    total_deal = float(deal_str) if deal_str else 0.0
                except ValueError:
                    total_deal = 0.0
                name = name.replace('&amp;', '&')[:200]
                db.session.add(Customer(
                    name=name, country=country, contact_person=contact_person,
                    email=email, phone=phone, address=address,
                    salesperson=salesperson, total_deal_usd=total_deal,
                ))
                count += 1
                if count % 500 == 0: db.session.commit()
            except Exception:
                pass
    db.session.commit()


def create_app():
    app = Flask(__name__)
    app.config['SECRET_KEY'] = 'pi-manager-secret-key-change-in-production'

    # Database — use env var DATABASE_DIR for Vercel, default to instance/
    db_dir = os.environ.get('DATABASE_DIR', os.path.join(APP_ROOT, 'instance'))
    db_path = os.path.join(db_dir, 'pi_manager.db')
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    app.config['TEMPLATES_AUTO_RELOAD'] = True

    # PDF output directory — use env var PDF_DIR for Vercel
    pdf_dir = os.environ.get('PDF_DIR', os.path.join(APP_ROOT, 'pdf'))
    app.config['PDF_DIR'] = pdf_dir
    upload_dir = os.environ.get('UPLOAD_DIR', os.path.join(APP_ROOT, 'static', 'uploads'))
    app.config['UPLOAD_DIR'] = upload_dir
    app.config['MAX_CONTENT_LENGTH'] = 256 * 1024 * 1024  # 256 MB max upload
    os.makedirs(pdf_dir, exist_ok=True)
    os.makedirs(upload_dir, exist_ok=True)
    os.makedirs(os.path.dirname(db_path), exist_ok=True)

    # Restore database from Vercel Blob if available
    blob_sync.init_db(db_path)

    db.init_app(app)

    with app.app_context():
        db.create_all()
        _migrate_db()

        # Seed default salespersons
        default_sp = [
            ('Shu Kei', '+86 17712333882', 'fantastic10086@gmail.com', '01075605503423165909'),
            ('Limon', '+86 15301506008', 'limon@qisuowelding.com', '0302320103321509442'),
            ('Kristi', '+86 18112500618', 'kristi@qisuowelding.com', '03236802374626412929'),
            ('Lu Yan', '+86 18019696608', 'luyan@qisuowelding.com', '0354364226451227898'),
            ('Yuna', '', '', ''),
            ('Delia', '', '', ''),
            ('Bai.', '', '', ''),
            ('Ma jilan', '', '', '036944123839016306'),
            ('QinQin', '18118336008', '', '2248391749772992006'),
        ]
        for name, phone, email, dt_uid in default_sp:
            if not Salesperson.query.filter_by(name=name).first():
                db.session.add(Salesperson(name=name, phone=phone, email=email, dingtalk_user_id=dt_uid))
        db.session.commit()

        # Seed default accounts (by name, won't overwrite existing)
        default_accounts = [
            ('克利斯达-农行', 'CHANGZHOU KLISTA INTERNATIONAL TRADE CO.LTD', 'AGRICULTURAL BANK OF CHINA H.O.BEIJING', '10618114040004700', 'ABOCCNBJ', 'klista', 'USD'),
            ('QISUO-花旗', 'Changzhou Q1 Suo Welding And Cutting Equipment Co., Ltd.', 'CITIBANK N. A. HONG KONG BRANCH', '39740000004173', 'CITIHKHXXXX', 'qisuo', 'USD'),
            ('姜舒棋的支付宝', 'CHANGZHOU KLISTA INTERNATIONAL TRADE CO.LTD', 'Alipay', '17712333882', '', 'klista', 'RMB'),
            ('姜舒棋的微信', 'CHANGZHOU KLISTA INTERNATIONAL TRADE CO.LTD', 'Wechat', '17712333882', '', 'klista', 'RMB'),
            ('戚所-阿里链接', 'Changzhou Q1 Suo Welding And Cutting Equipment Co., Ltd.', 'Alibaba', '', '', 'qisuo', 'USD'),
            ('克利斯达-阿里链接', 'CHANGZHOU KLISTA INTERNATIONAL TRADE CO.LTD', 'Alibaba', '', '', 'klista', 'USD'),
        ]
        for name, co_name, bank_name, acct_no, swift, brand, currency in default_accounts:
            if not Account.query.filter_by(name=name).first():
                db.session.add(Account(name=name, company_name=co_name, bank_name=bank_name, account_no=acct_no, swift_code=swift, brand=brand, currency=currency))
        db.session.commit()

        # Seed default customers from CSV (only if empty)
        if Customer.query.count() == 0:
            _seed_customers_from_csv()

        # Seed default products from CSV (only if empty)
        if Product.query.count() == 0:
            _seed_products_from_csv()

        # Seed default users
        default_users = [
            ('admin', 'admin123', 'admin', ''),
            ('shukei', 'admin123', 'admin', 'Shu Kei'),
            ('Ma Jilan', 'admin123', 'admin', ''),
            ('Limon', '123456', 'salesperson', 'Limon'),
            ('Kristi', '123456', 'salesperson', 'Kristi'),
            ('Lu Yan', '123456', 'salesperson', 'Lu Yan'),
            ('QinQin', '123456', 'salesperson', 'QinQin'),
        ]
        for uname, pwd, role, sp in default_users:
            if not User.query.filter_by(username=uname).first():
                db.session.add(User(
                    username=uname,
                    password_hash=generate_password_hash(pwd),
                    role=role,
                    salesperson_name=sp,
                ))
        db.session.commit()

    # Auto-sync database to Vercel Blob after write requests
    @app.after_request
    def _sync_to_blob(response):
        if request.method in ('POST', 'PUT', 'PATCH', 'DELETE'):
            if 200 <= response.status_code < 400:
                db_path = os.environ.get('DATABASE_DIR', os.path.join(APP_ROOT, 'instance'))
                db_path = os.path.join(db_path, 'pi_manager.db')
                blob_sync.sync_db_async(db_path)
        return response

    return app


app = create_app()


# ── Helper ─────────────────────────────────────────────────────────────
def _generate_pi_number():
    """Generate sequential PI number: PI-YYYYMMDD-NNN"""
    today = date.today()
    prefix = f"PI-{today.strftime('%Y%m%d')}-"
    # Find the max existing number instead of just counting (handles gaps from deletions)
    last = PI.query.filter(PI.pi_number.like(f'{prefix}%')).order_by(PI.pi_number.desc()).first()
    if last and last.pi_number.startswith(prefix):
        try:
            num = int(last.pi_number[len(prefix):]) + 1
        except ValueError:
            num = 1
    else:
        num = 1
    return f"{prefix}{num:03d}"

def _save_upload(file):
    """Save an uploaded file and return the filename."""
    if not file or file.filename == '':
        return ''
    filename = secure_filename(file.filename)
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ('.jpg', '.jpeg', '.png', '.gif', '.webp'):
        return ''
    unique_name = f"{uuid.uuid4().hex}{ext}"
    file.save(os.path.join(current_app.config['UPLOAD_DIR'], unique_name))
    return unique_name

ALLOWED_EXTENSIONS = {'jpg', 'jpeg', 'png', 'gif', 'webp'}


# ── Auth helpers ─────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def get_current_user():
    if 'user_id' in session:
        return User.query.get(session['user_id'])
    return None

def is_admin():
    return session.get('role') == 'admin'

def filter_by_user(query, model, salesperson_field='salesperson'):
    """Filter query by current user's salesperson if not admin."""
    if is_admin():
        return query
    sp = session.get('salesperson_name', '')
    if sp and hasattr(model, salesperson_field):
        return query.filter(getattr(model, salesperson_field) == sp)
    return query


# ═══════════════════════════════════════════════════════════════════════
#  ROUTES — Auth
# ═══════════════════════════════════════════════════════════════════════

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            if not user.active:
                flash('Account has been disabled. Contact admin.', 'danger')
                return render_template('login.html')
            session['user_id'] = user.id
            session['username'] = user.username
            session['role'] = user.role
            session['salesperson_name'] = user.salesperson_name
            flash(f'Welcome, {user.username}!', 'success')
            return redirect(url_for('index'))
        flash('Invalid username or password.', 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out.', 'info')
    return redirect(url_for('login'))


@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    user = get_current_user()
    sp = None
    if not is_admin() and user.salesperson_name:
        sp = Salesperson.query.filter_by(name=user.salesperson_name).first()
        # Auto-create salesperson record if missing
        if not sp:
            sp = Salesperson(name=user.salesperson_name)
            db.session.add(sp)
            db.session.commit()

    if request.method == 'POST':
        action = request.form.get('action', '')
        if action == 'change_password':
            current_pw = request.form.get('current_password', '').strip()
            new_pw = request.form.get('new_password', '').strip()
            confirm_pw = request.form.get('confirm_password', '').strip()
            if not current_pw or not new_pw:
                flash('All password fields are required.', 'danger')
            elif not check_password_hash(user.password_hash, current_pw):
                flash('Current password is incorrect.', 'danger')
            elif len(new_pw) < 4:
                flash('New password must be at least 4 characters.', 'danger')
            elif new_pw != confirm_pw:
                flash('New passwords do not match.', 'danger')
            else:
                user.password_hash = generate_password_hash(new_pw)
                db.session.commit()
                flash('Password changed successfully.', 'success')
            return redirect(url_for('profile'))
        elif sp:
            sp.phone = request.form.get('phone', '').strip()
            sp.email = request.form.get('email', '').strip()
            sp.notes = request.form.get('notes', '').strip()
            db.session.commit()
            flash('Profile updated.', 'success')
            return redirect(url_for('profile'))

    # Stats for salesperson
    stats = None
    pi_list = []
    if sp:
        stats = db.session.query(
            func.count(PI.id).label('pi_count'),
            func.coalesce(func.sum(PI.total_amount), 0).label('total_amount'),
        ).filter(PI.salesperson == sp.name).first()
        pi_list = PI.query.options(joinedload(PI.customer)).filter_by(salesperson=sp.name).order_by(PI.created_at.desc()).limit(10).all()

    return render_template('profile.html', sp=sp, stats=stats, pi_list=pi_list)


@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings_page():
    """Admin settings — DingTalk webhook & app config."""
    if not is_admin():
        flash('Admin only.', 'danger')
        return redirect(url_for('profile'))

    if request.method == 'POST':
        settings = _load_settings()
        import json as _json
        webhook = request.form.get('dingtalk_webhook', '').strip()
        appkey = request.form.get('dingtalk_appkey', '').strip()
        appsecret = request.form.get('dingtalk_appsecret', '').strip()
        agent_id = request.form.get('dingtalk_agent_id', '').strip()
        default_execs = request.form.getlist('default_executors')
        if webhook:
            settings['dingtalk_webhook'] = webhook
        if appkey:
            settings['dingtalk_appkey'] = appkey
        if appsecret:
            settings['dingtalk_appsecret'] = appsecret
        if agent_id:
            settings['dingtalk_agent_id'] = agent_id
        settings['dingtalk_default_executors'] = _json.dumps(default_execs)
        exchange_rate = request.form.get('exchange_rate', '').strip()
        if exchange_rate:
            settings['exchange_rate'] = exchange_rate
        _save_settings(settings)
        flash('Settings saved.', 'success')
        return redirect(url_for('settings_page'))

    settings = _load_settings()
    import json as _json
    # Get salespersons with DT user IDs for default executor selection
    dt_sps = Salesperson.query.filter(Salesperson.dingtalk_user_id != '').filter(Salesperson.dingtalk_user_id.isnot(None)).all()
    default_execs = _json.loads(settings.get('dingtalk_default_executors', '[]'))
    return render_template('settings.html',
                           dingtalk_webhook=settings.get('dingtalk_webhook', ''),
                           dingtalk_appkey=settings.get('dingtalk_appkey', ''),
                           dingtalk_appsecret=settings.get('dingtalk_appsecret', ''),
                           dingtalk_agent_id=settings.get('dingtalk_agent_id', ''),
                           exchange_rate=settings.get('exchange_rate', '7.0'),
                           webhook_ok=bool(settings.get('dingtalk_webhook')),
                           task_ok=bool(settings.get('dingtalk_appkey') and settings.get('dingtalk_appsecret')),
                           dt_sps=dt_sps,
                           default_execs=default_execs)


# ═══════════════════════════════════════════════════════════════════════
#  ROUTES — Dashboard
# ═══════════════════════════════════════════════════════════════════════

@app.route('/')
@login_required
def index():
    cust_q = filter_by_user(Customer.query, Customer, 'salesperson')
    customer_count = cust_q.count()
    product_count = Product.query.count()
    pi_q = filter_by_user(PI.query, PI, 'salesperson')
    pi_count = pi_q.count()
    recent_pis = pi_q.order_by(PI.created_at.desc()).limit(5).all()
    return render_template('index.html',
                           customer_count=customer_count,
                           product_count=product_count,
                           pi_count=pi_count,
                           recent_pis=recent_pis)


# ═══════════════════════════════════════════════════════════════════════
#  ROUTES — Customers
# ═══════════════════════════════════════════════════════════════════════

@app.route('/customers')
@login_required
def customer_list():
    search = request.args.get('search', '').strip()
    sort = request.args.get('sort', '').strip()
    sp_filter = request.args.get('sp', '').strip()
    deal_min_str = request.args.get('deal_min', '').strip()
    deal_max_str = request.args.get('deal_max', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()

    query = filter_by_user(Customer.query, Customer, 'salesperson')
    if search:
        query = query.filter(
            db.or_(
                Customer.name.ilike(f'%{search}%'),
                Customer.country.ilike(f'%{search}%'),
                Customer.contact_person.ilike(f'%{search}%'),
                Customer.email.ilike(f'%{search}%'),
            )
        )
    if sp_filter:
        query = query.filter(Customer.salesperson == sp_filter)
    if deal_min_str:
        try:
            query = query.filter(Customer.total_deal_usd >= float(deal_min_str))
        except ValueError:
            pass
    if deal_max_str:
        try:
            query = query.filter(Customer.total_deal_usd <= float(deal_max_str))
        except ValueError:
            pass
    if date_from:
        try:
            query = query.filter(Customer.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(Customer.created_at <= datetime.strptime(date_to, '%Y-%m-%d'))
        except ValueError:
            pass

    if sort == 'deal_desc':
        query = query.order_by(Customer.total_deal_usd.desc())
    elif sort == 'deal_asc':
        query = query.order_by(Customer.total_deal_usd.asc())
    else:
        query = query.order_by(Customer.created_at.desc())

    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = 50
    total = query.count()
    total_pages = max(1, (total + per_page - 1) // per_page)
    if page < 1: page = 1
    if page > total_pages: page = total_pages

    # Totals for current filter
    filter_total_deal = query.with_entities(func.coalesce(func.sum(Customer.total_deal_usd), 0)).scalar() or 0

    customers = query.limit(per_page).offset((page - 1) * per_page).all()
    return render_template('customers.html', customers=customers, search=search, sort=sort,
                           sp_filter=sp_filter, deal_min=deal_min_str, deal_max=deal_max_str,
                           date_from=date_from, date_to=date_to, page=page,
                           total_pages=total_pages, total=total,
                           filter_total_deal=filter_total_deal)


@app.route('/customers/add', methods=['GET', 'POST'])
@login_required
def customer_add():
    if request.method == 'POST':
        image_file = request.files.get('image')
        customer = Customer(
            name=request.form.get('name', '').strip(),
            country=request.form.get('country', '').strip(),
            contact_person=request.form.get('contact_person', '').strip(),
            email=request.form.get('email', '').strip(),
            phone=request.form.get('phone', '').strip(),
            address=request.form.get('address', '').strip(),
            salesperson=request.form.get('salesperson', '').strip(),
            image=_save_upload(image_file) if image_file else '',
            notes=request.form.get('notes', '').strip(),
        )
        if not customer.name:
            flash('Customer name is required.', 'danger')
            return render_template('customer_form.html', customer=customer, editing=False)
        if not customer.salesperson:
            flash('Please select a salesperson.', 'danger')
            return render_template('customer_form.html', customer=customer, editing=False)
        db.session.add(customer)
        db.session.commit()
        flash('Customer added successfully.', 'success')
        return redirect(url_for('customer_list'))
    return render_template('customer_form.html', customer=None, editing=False)


@app.route('/customers/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def customer_edit(id):
    customer = Customer.query.get_or_404(id)
    if request.method == 'POST':
        image_file = request.files.get('image')
        if image_file and image_file.filename:
            new_img = _save_upload(image_file)
            if new_img:
                # Delete old image
                if customer.image:
                    old_path = os.path.join(current_app.config['UPLOAD_DIR'], customer.image)
                    if os.path.exists(old_path):
                        os.remove(old_path)
                customer.image = new_img
        customer.name = request.form.get('name', '').strip()
        customer.country = request.form.get('country', '').strip()
        customer.contact_person = request.form.get('contact_person', '').strip()
        customer.email = request.form.get('email', '').strip()
        customer.phone = request.form.get('phone', '').strip()
        customer.address = request.form.get('address', '').strip()
        customer.salesperson = request.form.get('salesperson', '').strip()
        customer.notes = request.form.get('notes', '').strip()
        if not customer.name:
            flash('Customer name is required.', 'danger')
            return render_template('customer_form.html', customer=customer, editing=True)
        if not customer.salesperson:
            flash('Please select a salesperson.', 'danger')
            return render_template('customer_form.html', customer=customer, editing=True)
        db.session.commit()
        flash('Customer updated successfully.', 'success')
        return redirect(url_for('customer_list'))
    return render_template('customer_form.html', customer=customer, editing=True)


@app.route('/api/customers/add', methods=['POST'])
def api_customer_add():
    """AJAX endpoint — add a customer and return JSON."""
    name = request.form.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'error': 'Customer name is required.'}), 400

    customer = Customer(
        name=name,
        country=request.form.get('country', '').strip(),
        contact_person=request.form.get('contact_person', '').strip(),
        email=request.form.get('email', '').strip(),
        phone=request.form.get('phone', '').strip(),
        address=request.form.get('address', '').strip(),
        salesperson=request.form.get('salesperson', '').strip(),
        notes=request.form.get('notes', '').strip(),
    )
    db.session.add(customer)
    db.session.commit()

    return jsonify({
        'success': True,
        'customer': customer.to_dict(),
    })


@app.route('/api/customers/<int:id>/copy', methods=['POST'])
@login_required
def api_customer_copy(id):
    """Duplicate a customer with same info."""
    original = Customer.query.get_or_404(id)
    new_customer = Customer(
        name=original.name + ' (Copy)',
        country=original.country,
        contact_person=original.contact_person,
        email=original.email,
        phone=original.phone,
        address=original.address,
        salesperson=original.salesperson,
        notes=original.notes,
        image=original.image,
    )
    db.session.add(new_customer)
    db.session.commit()
    return jsonify({
        'success': True,
        'customer': new_customer.to_dict(),
    })


@app.route('/customers/<int:id>')
@login_required
def customer_detail(id):
    customer = Customer.query.get_or_404(id)
    pi_number = request.args.get('pi_number', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    product = request.args.get('product', '').strip()
    amount_min = request.args.get('amount_min', '').strip()
    amount_max = request.args.get('amount_max', '').strip()

    query = PI.query.options(
        joinedload(PI.items).joinedload(PIItem.product)
    ).filter_by(customer_id=id)

    if pi_number:
        query = query.filter(PI.pi_number.contains(pi_number))
    if date_from:
        try:
            d_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(PI.issue_date >= d_from)
        except ValueError:
            pass
    if date_to:
        try:
            d_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(PI.issue_date <= d_to)
        except ValueError:
            pass
    if amount_min:
        try:
            query = query.filter(PI.total_amount >= float(amount_min))
        except ValueError:
            pass
    if amount_max:
        try:
            query = query.filter(PI.total_amount <= float(amount_max))
        except ValueError:
            pass
    if product:
        query = query.join(PI.items).join(PIItem.product).filter(
            Product.name.ilike(f'%{product}%')
        ).distinct()

    pis = query.order_by(PI.created_at.desc()).all()
    resp = make_response(render_template('customer_detail.html', customer=customer, pis=pis,
                                          pi_number=pi_number, date_from=date_from, date_to=date_to,
                                          product=product, amount_min=amount_min, amount_max=amount_max))
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return resp


@app.route('/sales-stats')
@login_required
def sales_stats():
    """Payment received stats — paid PIs only, filterable by salesperson."""
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    preset = request.args.get('preset', '').strip()
    sp_filter = request.args.get('sp', '').strip()

    # Apply preset date ranges
    today = date.today()
    if preset == 'this_month':
        date_from = today.replace(day=1).strftime('%Y-%m-%d')
        date_to = today.strftime('%Y-%m-%d')
    elif preset == 'last_month':
        first_of_this_month = today.replace(day=1)
        last_of_last_month = first_of_this_month - timedelta(days=1)
        date_from = last_of_last_month.replace(day=1).strftime('%Y-%m-%d')
        date_to = last_of_last_month.strftime('%Y-%m-%d')
    elif preset == 'this_year':
        date_from = today.replace(month=1, day=1).strftime('%Y-%m-%d')
        date_to = today.strftime('%Y-%m-%d')

    paid_filter = (PI.received_amount > 0)  # include partial payments

    # Aggregate query — use received_amount (RMB→USD converted) for payment ranking
    amount_usd = func.sum(
        db.case(
            (PI.currency == 'RMB', PI.received_amount / 7.0),
            else_=PI.received_amount
        )
    )
    base_q = db.session.query(
        func.coalesce(db.func.nullif(PI.salesperson, ''), '(Unassigned)').label('salesperson'),
        func.count(PI.id).label('pi_count'),
        func.coalesce(amount_usd, 0).label('total_amount'),
        func.count(db.distinct(PI.customer_id)).label('customer_count'),
    ).filter(paid_filter)
    if sp_filter:
        base_q = base_q.filter(PI.salesperson == sp_filter)
    if date_from:
        try:
            base_q = base_q.filter(PI.issue_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to:
        try:
            base_q = base_q.filter(PI.issue_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass
    results = base_q.group_by(PI.salesperson).order_by(func.sum(PI.received_amount).desc()).all()

    grand_total = sum(r.total_amount for r in results)
    grand_pi_count = sum(r.pi_count for r in results)

    # Detailed PI list for the same filter (paid only) — eager load payments for fee display
    pi_detail_q = PI.query.options(
        joinedload(PI.customer),
        joinedload(PI.payments)
    ).filter(paid_filter).order_by(PI.created_at.desc())
    if sp_filter:
        pi_detail_q = pi_detail_q.filter(PI.salesperson == sp_filter)
    if date_from:
        try: pi_detail_q = pi_detail_q.filter(PI.issue_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
        except: pass
    if date_to:
        try: pi_detail_q = pi_detail_q.filter(PI.issue_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
        except: pass
    all_filtered_pis = pi_detail_q.all()

    # Check if any PI is in RMB
    has_rmb = any(pi.currency == 'RMB' for pi in all_filtered_pis)

    return render_template('sales_stats.html',
                           stats=results,
                           grand_total=grand_total,
                           grand_pi_count=grand_pi_count,
                           date_from=date_from,
                           date_to=date_to,
                           preset=preset,
                           sp_filter=sp_filter,
                           all_pis=all_filtered_pis,
                           has_rmb=has_rmb)


@app.route('/fees')
@login_required
def fees_report():
    """Fees report — PI total amount vs shipping cost breakdown."""
    # Only show PIs with payments (full or partial)
    pis = PI.query.options(joinedload(PI.customer)).filter(PI.received_amount > 0).order_by(PI.created_at.desc()).all()
    if not is_admin():
        user = get_current_user()
        sp = user.salesperson_name if user else ''
        pis = [p for p in pis if p.salesperson == sp]

    total_all = 0.0
    total_shipping = 0.0
    pi_list = []
    for p in pis:
        sym = '¥' if p.currency == 'RMB' else '$'
        sc = p.shipping_cost or 0.0
        subtotal = p.total_amount  # product subtotal only (no shipping)
        grand_total = p.total_amount + sc
        pi_list.append({
            'id': p.id, 'pi_number': p.pi_number,
            'customer': p.customer.name if p.customer else 'N/A',
            'country': p.customer.country if p.customer else '',
            'currency': p.currency, 'sym': sym,
            'total': grand_total, 'shipping': sc, 'subtotal': subtotal,
            'received': p.received_amount or 0.0,
            'date': p.issue_date.strftime('%Y-%m-%d') if p.issue_date else '',
            'paid': p.paid,
        })
        total_all += grand_total
        total_shipping += sc

    # Also fetch expenses grouped by PI
    all_expenses = Expense.query.order_by(Expense.created_at.desc()).all()
    exp_by_pi = {}
    for e in all_expenses:
        key = e.pi_id or 0
        if key not in exp_by_pi:
            exp_by_pi[key] = []
        exp_by_pi[key].append(e)
    total_expense = sum(e.amount for e in all_expenses)
    # Attach expenses to each PI in the list
    for p in pi_list:
        p['expenses'] = exp_by_pi.get(p['id'], [])
        p['exp_total'] = sum(e.amount for e in p['expenses'])

    resp = make_response(render_template('fees.html', pi_list=pi_list,
                           total_all=total_all, total_shipping=total_shipping,
                           product_total=total_all - total_shipping,
                           total_expense=total_expense))
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp


@app.route('/api/expenses', methods=['GET', 'POST'])
@login_required
def api_expenses():
    if request.method == 'POST':
        category = request.form.get('category', '').strip()
        amount_str = request.form.get('amount', '0').strip()
        currency = request.form.get('currency', 'USD').strip()
        note = request.form.get('note', '').strip()
        pi_id_str = request.form.get('pi_id', '').strip()
        try:
            pi_id = int(pi_id_str) if pi_id_str else None
        except ValueError:
            pi_id = None
        if not category:
            return jsonify({'success': False, 'error': 'Category is required.'}), 400
        try:
            amount = float(amount_str) if amount_str else 0.0
        except ValueError:
            return jsonify({'success': False, 'error': 'Invalid amount.'}), 400
        if amount <= 0:
            return jsonify({'success': False, 'error': 'Amount must be > 0.'}), 400
        exp = Expense(pi_id=pi_id, category=category, amount=amount, currency=currency, note=note)
        db.session.add(exp)
        db.session.commit()
        return jsonify({'success': True, 'expense': exp.to_dict()})
    # GET
    pi_id = request.args.get('pi_id', type=int)
    q = Expense.query
    if pi_id:
        q = q.filter_by(pi_id=pi_id)
    expenses = q.order_by(Expense.created_at.desc()).all()
    return jsonify([e.to_dict() for e in expenses])


@app.route('/api/expenses/<int:id>', methods=['DELETE'])
@login_required
def api_expense_delete(id):
    exp = Expense.query.get_or_404(id)
    db.session.delete(exp)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/customers/<int:id>/delete', methods=['POST'])
@login_required
def customer_delete(id):
    customer = Customer.query.get_or_404(id)
    db.session.delete(customer)
    db.session.commit()
    flash('Customer deleted successfully.', 'success')
    return redirect(url_for('customer_list'))


@app.route('/customers/<int:id>/reassign', methods=['POST'])
@login_required
def customer_reassign(id):
    if not is_admin():
        flash('Access denied.', 'danger')
        return redirect(url_for('customer_list'))
    customer = Customer.query.get_or_404(id)
    new_sp = request.form.get('salesperson', '').strip()
    old_sp = customer.salesperson
    if not new_sp:
        flash('Please select a salesperson.', 'danger')
        return redirect(url_for('customer_edit', id=id))
    # Update customer
    customer.salesperson = new_sp
    # Update all PIs under this customer
    count = PI.query.filter_by(customer_id=id).update({'salesperson': new_sp})
    db.session.commit()
    old_label = old_sp or 'unassigned'
    flash(f'Customer and {count} PI(s) reassigned from "{old_label}" to "{new_sp}".', 'success')
    return redirect(url_for('customer_list'))


# ═══════════════════════════════════════════════════════════════════════
#  ROUTES — Products
# ═══════════════════════════════════════════════════════════════════════

@app.route('/products')
@login_required
def product_list():
    search = request.args.get('search', '').strip()
    filter_type = request.args.get('filter', '').strip()
    sort = request.args.get('sort', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 50

    query = Product.query
    if search:
        query = query.filter(
            db.or_(
                Product.name.ilike(f'%{search}%'),
                Product.product_code.ilike(f'%{search}%'),
                Product.specification.ilike(f'%{search}%'),
                Product.chinese_name.ilike(f'%{search}%'),
            )
        )
    if filter_type == 'no_image':
        query = query.filter((Product.image == None) | (Product.image == ''))

    # Sorting
    sort_map = {
        'name_asc': Product.name.asc(),
        'name_desc': Product.name.desc(),
        'code_asc': Product.product_code.asc(),
        'code_desc': Product.product_code.desc(),
        'spec_asc': Product.specification.asc(),
        'spec_desc': Product.specification.desc(),
        'cn_asc': Product.chinese_name.asc(),
        'cn_desc': Product.chinese_name.desc(),
        'price_asc': Product.unit_price.asc(),
        'price_desc': Product.unit_price.desc(),
    }
    order_by = sort_map.get(sort, Product.created_at.desc())

    total = query.count()
    total_pages = max(1, (total + per_page - 1) // per_page)
    if page < 1: page = 1
    if page > total_pages: page = total_pages

    products = query.order_by(order_by).limit(per_page).offset((page - 1) * per_page).all()
    return render_template('products.html', products=products, search=search, filter=filter_type,
                           sort=sort, page=page, total_pages=total_pages, total=total)


@app.route('/products/import', methods=['GET', 'POST'])
@login_required
def product_import():
    """Import products from Excel file with floating images."""
    from openpyxl import load_workbook

    if request.method == 'POST':
        use_path = request.form.get('server_path', '').strip()
        own_tmp = False

        if use_path:
            # Import from server-side file path
            tmp_path = use_path
            if not os.path.exists(tmp_path):
                flash(f'File not found: {tmp_path}', 'danger')
                return redirect(url_for('product_import'))
            if not tmp_path.lower().endswith(('.xlsx', '.xlsm')):
                flash('File must be .xlsx or .xlsm.', 'danger')
                return redirect(url_for('product_import'))
            own_tmp = False
        else:
            # Uploaded file
            file = request.files.get('excel_file')
            if not file or not file.filename:
                flash('Please select an Excel file or enter a server path.', 'danger')
                return redirect(url_for('product_import'))

            if not file.filename.lower().endswith(('.xlsx', '.xlsm')):
                flash('Please upload a .xlsx or .xlsm file.', 'danger')
                return redirect(url_for('product_import'))

            tmp_path = os.path.join(app.config['UPLOAD_DIR'], f'_import_{uuid.uuid4().hex}.xlsx')
            file.save(tmp_path)
            own_tmp = True

        try:
            wb = load_workbook(tmp_path)
            ws = wb.active

            # ── Extract floating images and map to rows ──
            row_images = {}  # row -> list of image objects
            for img in ws._images:
                try:
                    anchor = img.anchor
                    # TwoCellAnchor or OneCellAnchor
                    if hasattr(anchor, '_from'):
                        row = anchor._from.row  # 0-indexed
                        col = anchor._from.col
                        if row not in row_images:
                            row_images[row] = []
                        row_images[row].append(img)
                except Exception:
                    pass

            # ── Parse data rows (starting from row 2, 0-indexed row 1) ──
            imported = 0
            with_images = 0
            skipped = 0

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False)):
                excel_row = row_idx + 1  # 0-indexed row number

                # Read cell values (columns A-E = indices 0-4)
                # A: No, B: Name, C: Code, D: Specification, E: Unit Price
                name = str(row[1].value).strip() if row[1].value else ''
                if not name or name == 'None':
                    skipped += 1
                    continue
                if len(name) > 200:
                    name = name[:200]

                code = str(row[2].value).strip() if len(row) > 2 and row[2].value else ''
                if len(code) > 100:
                    code = code[:100]

                spec = str(row[3].value).strip() if len(row) > 3 and row[3].value else ''
                if len(spec) > 200:
                    spec = spec[:200]

                # Parse price
                price = 0.0
                if len(row) > 4 and row[4].value is not None:
                    try:
                        price = float(str(row[4].value).strip().replace(',', ''))
                    except (ValueError, TypeError):
                        price = 0.0

                # ── Extract image for this row ──
                image_filename = ''
                if excel_row in row_images:
                    for img in row_images[excel_row]:
                        try:
                            # Save image data
                            img_data = img._data()
                            ext = img.format.lower() if img.format else 'png'
                            if ext in ('jpeg', 'jpg'):
                                ext = 'jpg'
                            elif ext not in ('png', 'gif', 'webp', 'bmp'):
                                ext = 'png'
                            unique_name = f"{uuid.uuid4().hex}.{ext}"
                            img_path = os.path.join(app.config['UPLOAD_DIR'], unique_name)
                            with open(img_path, 'wb') as f:
                                f.write(img_data)
                            image_filename = unique_name
                            with_images += 1
                            break  # take first image per row
                        except Exception:
                            pass

                # Check if product already exists (same name + code)
                existing = Product.query.filter_by(name=name, product_code=code).first()
                if existing:
                    # Update existing product
                    existing.specification = spec or existing.specification
                    existing.unit_price = price if price > 0 else existing.unit_price
                    if image_filename:
                        # Delete old image
                        if existing.image:
                            old_path = os.path.join(app.config['UPLOAD_DIR'], existing.image)
                            if os.path.exists(old_path):
                                os.remove(old_path)
                        existing.image = image_filename
                    imported += 1
                else:
                    product = Product(
                        name=name,
                        product_code=code,
                        specification=spec,
                        unit_price=price,
                        image=image_filename,
                    )
                    db.session.add(product)
                    imported += 1

            db.session.commit()
            flash(f'Imported {imported} products ({with_images} with images). {skipped} empty rows skipped.', 'success')

        except Exception as e:
            db.session.rollback()
            flash(f'Error reading Excel file: {str(e)}', 'danger')
        finally:
            # Clean up temp file (only if we created it)
            if own_tmp and os.path.exists(tmp_path):
                os.remove(tmp_path)

        return redirect(url_for('product_list'))

    return render_template('product_import.html')


@app.route('/products/add', methods=['GET', 'POST'])
@login_required
def product_add():
    if request.method == 'POST':
        try:
            unit_price = float(request.form.get('unit_price', '0').strip() or '0')
        except ValueError:
            unit_price = 0.0
        try:
            unit_price_rmb = float(request.form.get('unit_price_rmb', '0').strip() or '0')
        except ValueError:
            unit_price_rmb = 0.0
        image_file = request.files.get('image')
        image_filename = _save_upload(image_file) if image_file else ''
        product = Product(
            name=request.form.get('name', '').strip(),
            product_code=request.form.get('product_code', '').strip(),
            specification=request.form.get('specification', '').strip(),
            chinese_name=request.form.get('chinese_name', '').strip(),
            unit_price=unit_price,
            unit_price_rmb=unit_price_rmb,
            notes=request.form.get('notes', '').strip(),
            image=image_filename,
        )
        if not product.name:
            flash('Product name is required.', 'danger')
            return render_template('product_form.html', product=product, editing=False)
        db.session.add(product)
        db.session.commit()
        flash('Product added successfully.', 'success')
        return redirect(url_for('product_list'))
    return render_template('product_form.html', product=None, editing=False)


@app.route('/products/<int:id>/edit', methods=['GET', 'POST'])
def product_edit(id):
    product = Product.query.get_or_404(id)
    if request.method == 'POST':
        product.name = request.form.get('name', '').strip()
        product.product_code = request.form.get('product_code', '').strip()
        product.specification = request.form.get('specification', '').strip()
        product.chinese_name = request.form.get('chinese_name', '').strip()
        try:
            product.unit_price_rmb = float(request.form.get('unit_price_rmb', '0').strip() or '0')
        except ValueError:
            product.unit_price_rmb = 0.0
        try:
            product.unit_price = float(request.form.get('unit_price', '0').strip() or '0')
        except ValueError:
            product.unit_price = 0.0
        product.notes = request.form.get('notes', '').strip()
        image_file = request.files.get('image')
        if image_file and image_file.filename:
            new_img = _save_upload(image_file)
            if new_img:
                # Delete old image
                if product.image:
                    old_path = os.path.join(current_app.config['UPLOAD_DIR'], product.image)
                    if os.path.exists(old_path): os.remove(old_path)
                product.image = new_img
        if not product.name:
            flash('Product name is required.', 'danger')
            return render_template('product_form.html', product=product, editing=True)
        db.session.commit()
        flash('Product updated successfully.', 'success')
        return redirect(url_for('product_list'))
    return render_template('product_form.html', product=product, editing=True)


@app.route('/products/<int:id>/delete', methods=['POST'])
def product_delete(id):
    product = Product.query.get_or_404(id)
    if product.image:
        old_path = os.path.join(app.config['UPLOAD_DIR'], product.image)
        if os.path.exists(old_path):
            os.remove(old_path)
    db.session.delete(product)
    db.session.commit()
    flash('Product deleted.', 'success')
    return redirect(url_for('product_list'))


@app.route('/products/batch-delete', methods=['POST'])
@login_required
def product_batch_delete():
    """Batch delete selected products."""
    ids_str = request.form.get('ids', '')
    if not ids_str:
        flash('No products selected.', 'danger')
        return redirect(url_for('product_list'))
    ids = [int(i) for i in ids_str.split(',') if i.strip().isdigit()]
    if not ids:
        flash('No valid product IDs.', 'danger')
        return redirect(url_for('product_list'))
    for pid in ids:
        product = Product.query.get(pid)
        if product:
            if product.image:
                old_path = os.path.join(app.config['UPLOAD_DIR'], product.image)
                if os.path.exists(old_path):
                    os.remove(old_path)
            db.session.delete(product)
    db.session.commit()
    flash(f'{len(ids)} products deleted.', 'success')
    return redirect(url_for('product_list'))


@app.route('/api/products/add', methods=['POST'])
def api_product_add():
    """AJAX endpoint — add a product and return JSON."""
    try:
        unit_price = float(request.form.get('unit_price', '0').strip() or '0')
    except ValueError:
        unit_price = 0.0
    try:
        unit_price_rmb = float(request.form.get('unit_price_rmb', '0').strip() or '0')
    except ValueError:
        unit_price_rmb = 0.0

    image_file = request.files.get('image')
    product = Product(
        name=request.form.get('name', '').strip(),
        product_code=request.form.get('product_code', '').strip(),
        specification=request.form.get('specification', '').strip(),
        chinese_name=request.form.get('chinese_name', '').strip(),
        unit_price=unit_price,
        unit_price_rmb=unit_price_rmb,
        notes=request.form.get('notes', '').strip(),
        image=_save_upload(image_file) if image_file else '',
    )
    if not product.name:
        return jsonify({'success': False, 'error': 'Product name is required.'}), 400

    db.session.add(product)
    db.session.commit()

    return jsonify({
        'success': True,
        'product': product.to_dict(),
    })


@app.route('/api/products/search')
@login_required
def api_product_search():
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])
    pattern = f'%{q}%'
    products = Product.query.filter(
        db.or_(Product.name.ilike(pattern), Product.product_code.ilike(pattern))
    ).limit(50).all()
    return jsonify([{
        'id': p.id, 'name': p.name, 'chinese_name': p.chinese_name or '',
        'product_code': p.product_code or '', 'spec': p.specification or '',
        'price': p.unit_price, 'img': p.image or '',
    } for p in products])


@app.route('/api/products/<int:id>/update', methods=['POST'])
@login_required
def api_product_update(id):
    """Inline update product fields. Accepts JSON: {field: value}"""
    product = Product.query.get_or_404(id)
    data = request.get_json(silent=True) or {}
    allowed = {'name', 'product_code', 'specification', 'chinese_name', 'unit_price', 'unit_price_rmb'}
    for field, value in data.items():
        if field not in allowed:
            continue
        if field in ('unit_price', 'unit_price_rmb'):
            try:
                value = float(str(value).replace(',', ''))
            except (ValueError, TypeError):
                return jsonify({'success': False, 'error': 'Invalid price.'}), 400
        else:
            value = str(value).strip()
            if field == 'name' and not value:
                return jsonify({'success': False, 'error': 'Name required.'}), 400
        setattr(product, field, value)
    db.session.commit()
    return jsonify({
        'success': True,
        'product': product.to_dict(),
    })


@app.route('/api/products/<int:id>/copy', methods=['POST'])
@login_required
def api_product_copy(id):
    """Duplicate a product with same info."""
    original = Product.query.get_or_404(id)
    # Generate unique name and code to avoid conflicts
    new_name = original.name + ' (Copy)'
    new_code = original.product_code + '_copy' if original.product_code else 'copy'
    new_product = Product(
        name=new_name,
        product_code=new_code,
        specification=original.specification,
        chinese_name=original.chinese_name,
        unit_price=original.unit_price,
        unit_price_rmb=original.unit_price_rmb,
        notes=original.notes,
        image=original.image,  # share the same image file
    )
    db.session.add(new_product)
    db.session.commit()
    return jsonify({
        'success': True,
        'product': new_product.to_dict(),
    })


@app.route('/api/products/<int:id>/delete', methods=['POST'])
@login_required
def api_product_delete(id):
    """AJAX delete product."""
    product = Product.query.get_or_404(id)
    if product.image:
        old_path = os.path.join(app.config['UPLOAD_DIR'], product.image)
        if os.path.exists(old_path):
            os.remove(old_path)
    db.session.delete(product)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/translate', methods=['POST'])
@login_required
def api_translate():
    """Translate English text to Chinese using Google Translate API."""
    import urllib.request
    import json
    text = request.form.get('text', '').strip()
    if not text:
        return jsonify({'success': False, 'error': 'No text provided.'}), 400
    try:
        url = 'https://translate.googleapis.com/translate_a/single?client=gtx&sl=en&tl=zh-CN&dt=t&q=' + urllib.parse.quote(text)
        import ssl
        ctx = ssl._create_unverified_context()
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=5, context=ctx) as resp:
            data = json.loads(resp.read().decode('utf-8'))
        # Extract translated sentences from response
        result = ''.join([s[0] for s in data[0] if s[0]]) if data and data[0] else text
        return jsonify({'success': True, 'translated': result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_file(os.path.join(current_app.config['UPLOAD_DIR'], secure_filename(filename)))


# ═══════════════════════════════════════════════════════════════════════
#  ROUTES — Salespersons
# ═══════════════════════════════════════════════════════════════════════

@app.route('/salespersons')
@login_required
def salesperson_list():
    salespersons = Salesperson.query.order_by(Salesperson.name).all()
    # Compute customer counts per salesperson
    sp_stats = {}
    for sp in salespersons:
        total_cust = Customer.query.filter_by(salesperson=sp.name).count()
        paid_cust = db.session.query(func.count(func.distinct(PI.customer_id))).filter(
            PI.salesperson == sp.name,
            PI.paid == True
        ).scalar() or 0
        sp_stats[sp.name] = {'total_customers': total_cust, 'paid_customers': paid_cust}
    return render_template('salesperson_list.html', salespersons=salespersons, sp_stats=sp_stats)


@app.route('/salespersons/add', methods=['GET', 'POST'])
def salesperson_add():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('Salesperson name is required.', 'danger')
            return render_template('salesperson_form.html', sp=None, editing=False)
        if Salesperson.query.filter_by(name=name).first():
            flash('Salesperson name already exists.', 'danger')
            return render_template('salesperson_form.html', sp=None, editing=False)
        sp = Salesperson(
            name=name,
            phone=request.form.get('phone', '').strip(),
            email=request.form.get('email', '').strip(),
            dingtalk_user_id=request.form.get('dingtalk_user_id', '').strip(),
            notes=request.form.get('notes', '').strip(),
        )
        db.session.add(sp)
        db.session.commit()
        flash('Salesperson added successfully.', 'success')
        return redirect(url_for('salesperson_list'))
    return render_template('salesperson_form.html', sp=None, editing=False)


@app.route('/salespersons/<int:id>/edit', methods=['GET', 'POST'])
def salesperson_edit(id):
    sp = Salesperson.query.get_or_404(id)
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('Salesperson name is required.', 'danger')
            return render_template('salesperson_form.html', sp=sp, editing=True)
        existing = Salesperson.query.filter_by(name=name).first()
        if existing and existing.id != sp.id:
            flash('Salesperson name already exists.', 'danger')
            return render_template('salesperson_form.html', sp=sp, editing=True)
        sp.name = name
        sp.phone = request.form.get('phone', '').strip()
        sp.email = request.form.get('email', '').strip()
        sp.dingtalk_user_id = request.form.get('dingtalk_user_id', '').strip()
        sp.notes = request.form.get('notes', '').strip()
        db.session.commit()
        flash('Salesperson updated successfully.', 'success')
        return redirect(url_for('salesperson_list'))
    return render_template('salesperson_form.html', sp=sp, editing=True)


@app.route('/salespersons/<int:id>/delete', methods=['POST'])
def salesperson_delete(id):
    sp = Salesperson.query.get_or_404(id)
    db.session.delete(sp)
    db.session.commit()
    flash('Salesperson deleted successfully.', 'success')
    return redirect(url_for('salesperson_list'))


# ═══════════════════════════════════════════════════════════════════════
#  ROUTES — Users (Admin only)
# ═══════════════════════════════════════════════════════════════════════

@app.route('/users')
@login_required
def user_list():
    if not is_admin():
        flash('Access denied.', 'danger')
        return redirect(url_for('index'))
    users = User.query.order_by(User.created_at.desc()).all()
    return render_template('user_list.html', users=users)


@app.route('/users/add', methods=['GET', 'POST'])
@login_required
def user_add():
    if not is_admin():
        flash('Access denied.', 'danger')
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        role = request.form.get('role', 'salesperson').strip()
        salesperson_name = request.form.get('salesperson_name', '').strip()
        if not username or not password:
            flash('Username and password are required.', 'danger')
            return render_template('user_form.html', u=None, editing=False)
        if User.query.filter_by(username=username).first():
            flash('Username already exists.', 'danger')
            return render_template('user_form.html', u=None, editing=False)
        db.session.add(User(
            username=username,
            password_hash=generate_password_hash(password),
            role=role,
            salesperson_name=salesperson_name if role == 'salesperson' else '',
        ))
        db.session.commit()
        flash('User created.', 'success')
        return redirect(url_for('user_list'))
    return render_template('user_form.html', u=None, editing=False)


@app.route('/users/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def user_edit(id):
    if not is_admin():
        flash('Access denied.', 'danger')
        return redirect(url_for('index'))
    user = User.query.get_or_404(id)
    default_pw = 'admin123' if user.role == 'admin' else '123456'
    is_default = check_password_hash(user.password_hash, default_pw)
    pw_hint = f'Default: {default_pw}' if is_default else 'Custom (changed)'
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        role = request.form.get('role', 'salesperson').strip()
        salesperson_name = request.form.get('salesperson_name', '').strip()
        if not username:
            flash('Username is required.', 'danger')
            return render_template('user_form.html', u=user, editing=True, pw_hint=pw_hint)
        existing = User.query.filter_by(username=username).first()
        if existing and existing.id != user.id:
            flash('Username already exists.', 'danger')
            return render_template('user_form.html', u=user, editing=True, pw_hint=pw_hint)
        user.username = username
        if password:
            user.password_hash = generate_password_hash(password)
        user.role = role
        user.salesperson_name = salesperson_name if role == 'salesperson' else ''
        db.session.commit()
        flash('User updated.', 'success')
        return redirect(url_for('user_list'))
    return render_template('user_form.html', u=user, editing=True, pw_hint=pw_hint)


@app.route('/users/<int:id>/delete', methods=['POST'])
@login_required
def user_delete(id):
    if not is_admin():
        flash('Access denied.', 'danger')
        return redirect(url_for('index'))
    user = User.query.get_or_404(id)
    if user.username == 'admin':
        flash('Cannot delete the admin account.', 'danger')
        return redirect(url_for('user_list'))
    db.session.delete(user)
    db.session.commit()
    flash('User deleted.', 'success')
    return redirect(url_for('user_list'))


@app.route('/users/<int:id>/toggle-active', methods=['POST'])
@login_required
def user_toggle_active(id):
    if not is_admin():
        flash('Access denied.', 'danger')
        return redirect(url_for('index'))
    user = User.query.get_or_404(id)
    if user.username == 'admin':
        flash('Cannot disable the admin account.', 'danger')
        return redirect(url_for('user_list'))
    user.active = not user.active
    db.session.commit()
    status = 'enabled' if user.active else 'disabled'
    flash(f'User "{user.username}" {status}.', 'success')
    return redirect(url_for('user_list'))


@app.route('/users/<int:id>/reset-password', methods=['POST'])
@login_required
def user_reset_password(id):
    if not is_admin():
        flash('Access denied.', 'danger')
        return redirect(url_for('index'))
    user = User.query.get_or_404(id)
    new_pw = request.form.get('new_password', '').strip()
    if not new_pw or len(new_pw) < 4:
        flash('Password must be at least 4 characters.', 'danger')
        return redirect(url_for('user_list'))
    user.password_hash = generate_password_hash(new_pw)
    db.session.commit()
    flash(f'Password for "{user.username}" has been reset.', 'success')
    return redirect(url_for('user_list'))


# ═══════════════════════════════════════════════════════════════════════
#  ROUTES — Accounts (Admin only)
# ═══════════════════════════════════════════════════════════════════════

@app.route('/accounts')
@login_required
def account_list():
    if not is_admin(): return redirect(url_for('index'))
    accounts = Account.query.order_by(Account.name).all()
    return render_template('account_list.html', accounts=accounts)

@app.route('/accounts/add', methods=['GET', 'POST'])
@login_required
def account_add():
    if not is_admin(): return redirect(url_for('index'))
    if request.method == 'POST':
        db.session.add(Account(
            name=request.form.get('name','').strip(),
            company_name=request.form.get('company_name','').strip(),
            bank_name=request.form.get('bank_name','').strip(),
            account_no=request.form.get('account_no','').strip(),
            swift_code=request.form.get('swift_code','').strip(),
            brand=request.form.get('brand','klista').strip(),
            currency=request.form.get('currency','USD').strip(),
            notes=request.form.get('notes','').strip(),
        ))
        db.session.commit()
        flash('Account added.', 'success')
        return redirect(url_for('account_list'))
    return render_template('account_form.html', a=None, editing=False)

@app.route('/accounts/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def account_edit(id):
    if not is_admin(): return redirect(url_for('index'))
    a = Account.query.get_or_404(id)
    if request.method == 'POST':
        a.name = request.form.get('name','').strip()
        a.company_name = request.form.get('company_name','').strip()
        a.bank_name = request.form.get('bank_name','').strip()
        a.account_no = request.form.get('account_no','').strip()
        a.swift_code = request.form.get('swift_code','').strip()
        a.brand = request.form.get('brand','klista').strip()
        a.currency = request.form.get('currency','USD').strip()
        a.notes = request.form.get('notes','').strip()
        db.session.commit()
        flash('Account updated.', 'success')
        return redirect(url_for('account_list'))
    return render_template('account_form.html', a=a, editing=True)

@app.route('/accounts/<int:id>/delete', methods=['POST'])
@login_required
def account_delete(id):
    if not is_admin(): return redirect(url_for('index'))
    db.session.delete(Account.query.get_or_404(id))
    db.session.commit()
    flash('Account deleted.', 'success')
    return redirect(url_for('account_list'))


# ═══════════════════════════════════════════════════════════════════════
#  ROUTES — Suppliers
# ═══════════════════════════════════════════════════════════════════════

@app.route('/suppliers')
@login_required
def supplier_list():
    if not is_admin(): return redirect(url_for('index'))
    suppliers = Supplier.query.order_by(Supplier.name).all()
    return render_template('supplier_list.html', suppliers=suppliers)

@app.route('/suppliers/<int:id>')
@login_required
def supplier_detail(id):
    """Supplier detail page with procurement history and outstanding balance."""
    if not is_admin(): return redirect(url_for('index'))
    s = Supplier.query.get_or_404(id)
    # All procurement items for this supplier, with PI info
    procs = Procurement.query.options(
        db.joinedload(Procurement.pi),
        db.joinedload(Procurement.pi_item).joinedload(PIItem.product)
    ).filter_by(supplier_id=id).order_by(Procurement.created_at.desc()).all()

    # Calculate totals
    total_proc = sum(p.total for p in procs)
    # Group by PI
    pi_groups = {}
    for p in procs:
        pi_id = p.pi_id
        if pi_id not in pi_groups:
            pi_groups[pi_id] = {'pi': p.pi, 'proc_items': [], 'subtotal': 0}
        pi_groups[pi_id]['proc_items'].append(p)
        pi_groups[pi_id]['subtotal'] += p.total

    return render_template('supplier_detail.html', supplier=s, procs=procs,
                           total_proc=total_proc, pi_groups=pi_groups,
                           count=len(procs))


@app.route('/suppliers/<int:id>/statement')
@login_required
def supplier_statement(id):
    """Generate and download supplier statement PDF."""
    if not is_admin(): return redirect(url_for('index'))
    s = Supplier.query.get_or_404(id)
    procs = Procurement.query.options(
        db.joinedload(Procurement.pi),
        db.joinedload(Procurement.pi_item).joinedload(PIItem.product)
    ).filter_by(supplier_id=id).order_by(Procurement.created_at.asc()).all()
    total_proc = sum(p.total for p in procs)

    pdf_dir = os.path.join(APP_ROOT, 'pdf')
    os.makedirs(pdf_dir, exist_ok=True)
    safe_name = secure_filename(s.name)
    filename = f'statement_{safe_name}_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.pdf'
    filepath = os.path.join(pdf_dir, filename)

    generate_supplier_statement(s, procs, total_proc, filepath)

    return send_file(
        filepath, as_attachment=True,
        download_name=f'对账单_{s.name}_{datetime.utcnow().strftime("%Y%m%d")}.pdf',
        mimetype='application/pdf'
    )


@app.route('/suppliers/add', methods=['GET', 'POST'])
@login_required
def supplier_add():
    if not is_admin(): return redirect(url_for('index'))
    if request.method == 'POST':
        s = Supplier(
            name=request.form.get('name','').strip(),
            contact_person=request.form.get('contact_person','').strip(),
            phone=request.form.get('phone','').strip(),
            email=request.form.get('email','').strip(),
            address=request.form.get('address','').strip(),
            notes=request.form.get('notes','').strip(),
        )
        if not s.name:
            flash('Supplier name is required.', 'danger')
            return render_template('supplier_form.html', s=s, editing=False)
        db.session.add(s)
        db.session.commit()
        flash('Supplier added.', 'success')
        return redirect(url_for('supplier_list'))
    return render_template('supplier_form.html', s=None, editing=False)

@app.route('/suppliers/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def supplier_edit(id):
    if not is_admin(): return redirect(url_for('index'))
    s = Supplier.query.get_or_404(id)
    if request.method == 'POST':
        s.name = request.form.get('name','').strip()
        s.contact_person = request.form.get('contact_person','').strip()
        s.phone = request.form.get('phone','').strip()
        s.email = request.form.get('email','').strip()
        s.address = request.form.get('address','').strip()
        s.notes = request.form.get('notes','').strip()
        if not s.name:
            flash('Supplier name is required.', 'danger')
            return render_template('supplier_form.html', s=s, editing=True)
        db.session.commit()
        flash('Supplier updated.', 'success')
        return redirect(url_for('supplier_list'))
    return render_template('supplier_form.html', s=s, editing=True)

@app.route('/suppliers/<int:id>/delete', methods=['POST'])
@login_required
def supplier_delete(id):
    if not is_admin(): return redirect(url_for('index'))
    db.session.delete(Supplier.query.get_or_404(id))
    db.session.commit()
    flash('Supplier deleted.', 'success')
    return redirect(url_for('supplier_list'))

@app.route('/api/suppliers')
@login_required
def api_supplier_list():
    suppliers = Supplier.query.order_by(Supplier.name).all()
    return jsonify([s.to_dict() for s in suppliers])


# ═══════════════════════════════════════════════════════════════════════
#  ROUTES — PI (Proforma Invoice)
# ═══════════════════════════════════════════════════════════════════════

@app.route('/pi/create', methods=['GET', 'POST'])
@login_required
def pi_create():
    customers = filter_by_user(Customer.query, Customer, 'salesperson').order_by(Customer.name).all()
    products = Product.query.order_by(Product.name).all()
    all_salespersons = Salesperson.query.order_by(Salesperson.name).all()
    all_accounts = Account.query.order_by(Account.name).all()
    admin_flag = is_admin()

    if request.method == 'POST':
        customer_id = request.form.get('customer_id', type=int)
        salesperson = request.form.get('salesperson', '').strip()
        payment_terms = request.form.get('payment_terms', '').strip()
        bank_info = request.form.get('bank_info', '').strip()
        notes = request.form.get('notes', '').strip()
        issue_date_str = request.form.get('issue_date', '').strip()
        currency = request.form.get('currency', 'USD').strip()
        company = request.form.get('company', 'klista').strip()
        account_id = request.form.get('account_id', type=int)

        if not customer_id:
            flash('Please select a customer.', 'danger')
            return render_template('create_pi.html', customers=customers,
                                   products=products, pi=None,
                                   today=date.today().strftime('%Y-%m-%d'),
                                   selected_customer_id=None,
                                   all_salespersons=all_salespersons,
                                   all_accounts=all_accounts,
                                   is_admin=admin_flag)

        if not salesperson:
            flash('Please select a salesperson.', 'danger')
            return render_template('create_pi.html', customers=customers,
                                   products=products, pi=None,
                                   today=date.today().strftime('%Y-%m-%d'),
                                   selected_customer_id=customer_id,
                                   all_salespersons=all_salespersons,
                                   all_accounts=all_accounts,
                                   is_admin=admin_flag)

        shipping_address = request.form.get('shipping_address', '').strip()
        try:
            shipping_cost = float(request.form.get('shipping_cost', '0').strip() or '0')
        except ValueError:
            shipping_cost = 0.0

        # Parse issue date
        try:
            issue_date = datetime.strptime(issue_date_str, '%Y-%m-%d').date() if issue_date_str else date.today()
        except ValueError:
            issue_date = date.today()

        # Collect selected products with quantities
        selected_items = []
        total_amount = 0.0
        for product in products:
            qty_key = f'qty_{product.id}'
            selected_key = f'selected_{product.id}'
            if request.form.get(selected_key) == 'on':
                try:
                    qty = int(request.form.get(qty_key, '1').strip() or '1')
                except ValueError:
                    qty = 1
                unit_price_str = request.form.get(f'unit_price_{product.id}', '').strip()
                try:
                    unit_price = float(unit_price_str) if unit_price_str else product.unit_price
                except ValueError:
                    unit_price = product.unit_price
                if qty > 0:
                    amount = round(unit_price * qty, 2)
                    selected_items.append({
                        'product': product,
                        'quantity': qty,
                        'unit_price': unit_price,
                        'amount': amount,
                    })
                    total_amount += amount

        if not selected_items:
            flash('Please select at least one product.', 'danger')
            return render_template('create_pi.html', customers=customers,
                                   products=products, pi=None,
                                   today=issue_date.strftime('%Y-%m-%d'),
                                   selected_customer_id=customer_id,
                                   all_salespersons=all_salespersons,
                                   all_accounts=all_accounts,
                                   is_admin=admin_flag)

        total_amount = round(total_amount, 2)

        # Create PI record
        pi_number = _generate_pi_number()
        bank_info_from_account = ''
        if account_id:
            acct = Account.query.get(account_id)
            if acct: bank_info_from_account = acct.bank_info()

        # total_amount = product subtotal only (shipping stored separately)
        total_amount = round(total_amount, 2)

        pi = PI(
            pi_number=pi_number,
            customer_id=customer_id,
            issue_date=issue_date,
            payment_terms=payment_terms or '100% TT before shipment',
            bank_info=bank_info or bank_info_from_account,
            salesperson=salesperson,
            currency=currency,
            company=company,
            total_amount=total_amount,
            shipping_cost=shipping_cost,
            shipping_note=request.form.get('shipping_note', '').strip(),
            shipping_address=shipping_address,
            notes=notes,
        )
        db.session.add(pi)
        db.session.flush()  # Get pi.id

        # Create PI items
        for item in selected_items:
            pi_item = PIItem(
                pi_id=pi.id,
                product_id=item['product'].id,
                quantity=item['quantity'],
                unit_price=item['unit_price'],
                amount=item['amount'],
            )
            db.session.add(pi_item)

        db.session.flush()

        # Generate PDF
        try:
            # Reload PI with relationships for PDF generation
            pi_full = PI.query.options(
                joinedload(PI.items).joinedload(PIItem.product)
            ).get(pi.id)
            sp = Salesperson.query.filter_by(name=pi.salesperson).first()
            sp_info = {'phone': sp.phone, 'email': sp.email} if sp else {}
            pdf_filename = generate_pi_pdf(pi_full, app.config['PDF_DIR'], sp_info)
            excel_filename = generate_pi_excel(pi_full, app.config['PDF_DIR'])
            pi.pdf_path = pdf_filename
            pi.excel_path = excel_filename
        except Exception as e:
            db.session.rollback()
            flash(f'Error generating PDF: {str(e)}', 'danger')
            return render_template('create_pi.html', customers=customers,
                                   products=products, pi=None,
                                   today=issue_date.strftime('%Y-%m-%d'),
                                   selected_customer_id=customer_id,
                                   all_salespersons=all_salespersons,
                                   all_accounts=all_accounts,
                                   is_admin=admin_flag)

        db.session.commit()
        flash(f'PI {pi_number} created and PDF generated.', 'success')
        return redirect(url_for('pi_list'))

    # GET request
    selected_customer_id = request.args.get('customer_id', type=int)
    preselected_salesperson = ''
    if selected_customer_id:
        cust = Customer.query.get(selected_customer_id)
        if cust:
            preselected_salesperson = cust.salesperson
    all_salespersons = Salesperson.query.order_by(Salesperson.name).all()
    all_accounts = Account.query.order_by(Account.name).all()
    return render_template('create_pi.html', customers=customers,
                           products=products, pi=None,
                           today=date.today().strftime('%Y-%m-%d'),
                           selected_customer_id=selected_customer_id,
                           preselected_salesperson=preselected_salesperson,
                           all_salespersons=all_salespersons,
                           all_accounts=all_accounts,
                           is_admin=is_admin())


@app.route('/pi/list')
@login_required
def pi_list():
    salesperson_filter = request.args.get('salesperson', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()

    query = filter_by_user(PI.query.options(joinedload(PI.customer)), PI, 'salesperson')

    if salesperson_filter:
        query = query.filter(PI.salesperson == salesperson_filter)
    if date_from:
        try:
            d_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(PI.issue_date >= d_from)
        except ValueError:
            pass
    if date_to:
        try:
            d_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(PI.issue_date <= d_to)
        except ValueError:
            pass

    pis = query.order_by(PI.created_at.desc()).all()

    # Total for this filtered set
    filtered_total = sum(pi.total_amount for pi in pis)

    resp = make_response(render_template('pi_list.html', pis=pis,
                           salesperson_filter=salesperson_filter,
                           date_from=date_from, date_to=date_to,
                           filtered_total=filtered_total))
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp


@app.route('/pi/<int:id>')
@login_required
def pi_detail(id):
    pi = PI.query.options(
        joinedload(PI.customer),
        joinedload(PI.items).joinedload(PIItem.product),
    ).get_or_404(id)
    resp = make_response(render_template('pi_detail.html', pi=pi))
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp


@app.route('/pi/<int:id>/preview')
@login_required
def pi_preview(id):
    pi = PI.query.options(
        joinedload(PI.customer),
        joinedload(PI.items).joinedload(PIItem.product),
    ).get_or_404(id)
    return render_template('pi_preview.html', pi=pi)


@app.route('/pi/<int:id>/download')
def pi_download(id):
    pi = PI.query.get_or_404(id)
    if not pi.pdf_path:
        flash('PDF file not found.', 'danger')
        return redirect(url_for('pi_detail', id=id))
    filepath = os.path.join(app.config['PDF_DIR'], pi.pdf_path)
    if not os.path.exists(filepath):
        flash('PDF file does not exist on disk.', 'danger')
        return redirect(url_for('pi_detail', id=id))
    return send_file(filepath, as_attachment=True, download_name=pi.pdf_path)


@app.route('/pi/<int:id>/excel')
@login_required
def pi_excel_download(id):
    pi = PI.query.options(
        joinedload(PI.customer),
        joinedload(PI.items).joinedload(PIItem.product),
    ).get_or_404(id)
    # Auto-generate if missing
    if not pi.excel_path or not os.path.exists(os.path.join(app.config['PDF_DIR'], pi.excel_path)):
        try:
            fn = generate_pi_excel(pi, app.config['PDF_DIR'])
            pi.excel_path = fn
            db.session.commit()
        except Exception as e:
            flash(f'Error generating Excel: {e}', 'danger')
            return redirect(url_for('pi_detail', id=id))
    filepath = os.path.join(app.config['PDF_DIR'], pi.excel_path)
    return send_file(filepath, as_attachment=True, download_name=pi.excel_path)


@app.route('/pi/<int:id>/toggle-paid', methods=['POST'])
@login_required
def pi_toggle_paid(id):
    pi = PI.query.get_or_404(id)
    received_str = request.form.get('received_amount', '').strip()
    fee_str = request.form.get('fee', '0').strip()
    order_no = request.form.get('order_no', '').strip()
    try:
        amount = float(received_str) if received_str else pi.total_amount
    except ValueError:
        amount = pi.total_amount
    try:
        fee = float(fee_str) if fee_str else 0.0
    except ValueError:
        fee = 0.0
    if amount <= 0:
        flash('Amount must be greater than 0.', 'danger')
        return redirect(request.referrer or url_for('pi_list'))
    if not order_no:
        flash('Order number is required. (Alibaba Trade Assurance / Alipay / WeChat order number)', 'danger')
        return redirect(request.referrer or url_for('pi_list'))

    # Prevent duplicate submission (same order_no for same PI)
    existing = Payment.query.filter_by(pi_id=pi.id, order_no=order_no).first()
    if existing:
        flash('This order number has already been recorded for this PI.', 'warning')
        return redirect(request.referrer or url_for('pi_list'))

    # Check total received won't exceed PI grand total (product + shipping)
    pi_grand_total = pi.total_amount + (pi.shipping_cost or 0)
    current_received = sum(p.amount for p in pi.payments)
    if current_received + amount > pi_grand_total:
        sym = '¥' if pi.currency == 'RMB' else '$'
        remaining = pi_grand_total - current_received
        flash(f'Payment exceeds PI total. Remaining: {sym}{remaining:,.2f} (PI total: {sym}{pi_grand_total:,.2f})', 'danger')
        return redirect(request.referrer or url_for('pi_list'))

    # Add payment record
    db.session.add(Payment(pi_id=pi.id, amount=amount, fee=fee, order_no=order_no))
    db.session.flush()
    # Expire the payments relationship so it reloads with the new record
    db.session.expire(pi, ['payments'])

    # Update PI aggregates (amount + fee counts toward paid, but only amount toward deal)
    total_paid = sum(p.amount + (p.fee or 0) for p in pi.payments)
    pi.received_amount = sum(p.amount for p in pi.payments)
    pi_grand_total = pi.total_amount + (pi.shipping_cost or 0)
    pi.paid = total_paid >= pi_grand_total

    # Update customer's total_deal_usd (excludes fee)
    customer = Customer.query.get(pi.customer_id)
    if customer:
        paid_pis = PI.query.filter_by(customer_id=customer.id, paid=True).all()
        total = 0.0
        for p in paid_pis:
            deal_amount = p.received_amount if p.received_amount > 0 else p.total_amount
            if p.currency == 'RMB':
                deal_amount = round(deal_amount / 7.0, 2)
            total += deal_amount
        customer.total_deal_usd = round(total, 2)

    db.session.commit()

    # ── DingTalk 喜报 ──
    sym = '¥' if (pi.currency == 'RMB') else '$'
    cust_name = pi.customer.name if pi.customer else 'N/A'
    sp_name = pi.salesperson or 'N/A'
    status = '🎉 全额付清' if pi.paid else ('📥 部分付款' if pi.received_amount > 0 else '📝 首笔付款')
    ding_title = f'💰 回款喜报 — {cust_name}'
    ding_text = (
        f'## 💰 回款喜报\n\n'
        f'**客户：** {cust_name}\n\n'
        f'**PI单号：** {pi.pi_number}\n\n'
        f'**业务员：** {sp_name}\n\n'
        f'**本次收款：** {sym}{amount:,.2f}\n\n'
        f'**手续费：** {sym}{fee:,.2f}\n\n'
        f'**订单编号：** {order_no}\n\n'
        f'**累计已收：** {sym}{pi.received_amount:,.2f} / {sym}{pi.total_amount:,.2f}\n\n'
        f'**状态：** {status}\n\n'
        f'> {COMPANY_CONFIG["name"]}  \n'
        f'> {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    )
    _send_dingtalk(ding_title, ding_text)

    # ── DingTalk 待办 ──
    sp = Salesperson.query.filter_by(name=sp_name).first()
    dt_user_id = sp.dingtalk_user_id if sp else ''
    if dt_user_id:
        task_subject = f'{sp_name} - {cust_name} - {sym}{pi.received_amount:,.2f}/{sym}{pi.total_amount:,.2f} {status}'
        task_desc = (
            f'**客户：** {cust_name}\n'
            f'**PI单号：** {pi.pi_number}\n'
            f'**业务员：** {sp_name}\n'
            f'**本次收款：** {sym}{amount:,.2f}（手续费 {sym}{fee:,.2f}）\n'
            f'**订单编号：** {order_no}\n'
            f'**累计已收：** {sym}{pi.received_amount:,.2f} / {sym}{pi.total_amount:,.2f}\n'
            f'**状态：** {status}\n'
        )
        if pi.notes:
            task_desc += f'\n**备注：** {pi.notes}\n'

        # Build executor list: salesperson first, then default executors
        import json as _json
        settings = _load_settings()
        default_execs = _json.loads(settings.get('dingtalk_default_executors', '[]'))
        executor_ids = [dt_user_id]
        for eid in default_execs:
            if eid != dt_user_id:
                executor_ids.append(eid)

        pdf_path = os.path.join(app.config['PDF_DIR'], pi.pdf_path) if pi.pdf_path else ''
        task_id = _dingtalk_create_task(executor_ids, task_subject, task_desc)

    remaining = max(0, pi.total_amount - total_paid)
    flash(f'Payment ${amount:,.2f} (+fee ${fee:,.2f}) recorded. Remaining: ${remaining:,.2f}.', 'success')
    return redirect(request.referrer or url_for('pi_list'))

@app.route('/pi/<int:id>/payment/<int:pid>/delete', methods=['POST'])
@login_required
def payment_delete(id, pid):
    payment = Payment.query.get_or_404(pid)
    pi = PI.query.get_or_404(id)
    db.session.delete(payment)
    db.session.flush()
    db.session.expire(pi, ['payments'])
    # Recalculate
    total_paid = sum(p.amount + (p.fee or 0) for p in pi.payments)
    pi.received_amount = sum(p.amount for p in pi.payments)
    pi_grand_total = pi.total_amount + (pi.shipping_cost or 0)
    pi.paid = total_paid >= pi_grand_total
    # Update customer
    customer = Customer.query.get(pi.customer_id)
    if customer:
        paid_pis = PI.query.filter_by(customer_id=customer.id, paid=True).all()
        total = 0.0
        for p in paid_pis:
            deal_amount = p.received_amount if p.received_amount > 0 else p.total_amount
            if p.currency == 'RMB': deal_amount = round(deal_amount / 7.0, 2)
            total += deal_amount
        customer.total_deal_usd = round(total, 2)
    db.session.commit()
    flash('Payment record deleted.', 'info')
    return redirect(request.referrer or url_for('pi_list'))


@app.route('/api/pi/<int:id>/payments')
@login_required
def api_pi_payments(id):
    pi = PI.query.get_or_404(id)
    grand_total = pi.total_amount + (pi.shipping_cost or 0)
    return jsonify({
        'currency': pi.currency or 'USD',
        'total_amount': pi.total_amount,
        'grand_total': grand_total,
        'received': pi.received_amount or 0,
        'paid': pi.paid,
        'payments': [{'id': p.id, 'amount': p.amount, 'fee': p.fee or 0, 'order_no': p.order_no or '', 'created_at': p.created_at.strftime('%Y-%m-%d %H:%M') if p.created_at else ''} for p in pi.payments]
    })


@app.route('/pi/<int:id>/delete', methods=['POST'])
def pi_delete(id):
    pi = PI.query.get_or_404(id)
    # Delete PDF file if exists
    if pi.pdf_path:
        filepath = os.path.join(app.config['PDF_DIR'], pi.pdf_path)
        if os.path.exists(filepath):
            os.remove(filepath)
    db.session.delete(pi)
    db.session.commit()
    flash('PI deleted successfully.', 'success')
    return redirect(url_for('pi_list'))


@app.route('/pi/<int:id>/copy', methods=['POST'])
@login_required
def pi_copy(id):
    """Copy a PI: same info, new date and PI number."""
    original = PI.query.options(
        joinedload(PI.items).joinedload(PIItem.product)
    ).get_or_404(id)

    # Generate new PI number with today's date
    pi_number = _generate_pi_number()

    # Create new PI with same data
    new_pi = PI(
        pi_number=pi_number,
        customer_id=original.customer_id,
        issue_date=date.today(),
        payment_terms=original.payment_terms,
        bank_info=original.bank_info,
        salesperson=original.salesperson,
        currency=original.currency or 'USD',
        company=original.company or 'klista',
        total_amount=original.total_amount,
        shipping_cost=original.shipping_cost or 0.0,
        shipping_note=original.shipping_note or '',
        shipping_address=original.shipping_address or '',
        notes=original.notes,
    )
    db.session.add(new_pi)
    db.session.flush()  # Get new_pi.id

    # Copy all PI items
    for item in original.items:
        new_item = PIItem(
            pi_id=new_pi.id,
            product_id=item.product_id,
            quantity=item.quantity,
            unit_price=item.unit_price,
            amount=item.amount,
        )
        db.session.add(new_item)

    db.session.flush()

    # Generate PDF and Excel
    try:
        pi_full = PI.query.options(
            joinedload(PI.items).joinedload(PIItem.product)
        ).get(new_pi.id)
        sp = Salesperson.query.filter_by(name=new_pi.salesperson).first()
        sp_info = {'phone': sp.phone, 'email': sp.email} if sp else {}
        pdf_filename = generate_pi_pdf(pi_full, app.config['PDF_DIR'], sp_info)
        excel_filename = generate_pi_excel(pi_full, app.config['PDF_DIR'])
        new_pi.pdf_path = pdf_filename
        new_pi.excel_path = excel_filename
    except Exception as e:
        db.session.rollback()
        flash(f'Error generating PDF: {str(e)}', 'danger')
        return redirect(url_for('pi_list'))

    db.session.commit()
    flash(f'PI {pi_number} copied from {original.pi_number}. You can now edit it.', 'success')
    return redirect(url_for('pi_edit', id=new_pi.id))


def _preload_products(pi):
    preload = {}
    for item in pi.items:
        pid = str(item.product_id)
        prod = item.product
        if prod:
            preload[pid] = {'id': prod.id, 'name': prod.name, 'code': prod.product_code,
                            'spec': prod.specification or '', 'price': item.unit_price,
                            'img': prod.image or '', 'qty': item.quantity}
    return preload


@app.route('/pi/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def pi_edit(id):
    pi = PI.query.options(
        joinedload(PI.customer),
        joinedload(PI.items).joinedload(PIItem.product),
    ).get_or_404(id)
    customers = Customer.query.order_by(Customer.name).all()
    products = Product.query.order_by(Product.name).all()
    # Map existing items: product_id -> {quantity, selected}
    existing_items = {item.product_id: item.quantity for item in pi.items}
    preload = _preload_products(pi)

    if request.method == 'POST':
        customer_id = request.form.get('customer_id', type=int)
        salesperson = request.form.get('salesperson', '').strip()
        payment_terms = request.form.get('payment_terms', '').strip()
        bank_info = request.form.get('bank_info', '').strip()
        notes = request.form.get('notes', '').strip()
        issue_date_str = request.form.get('issue_date', '').strip()
        currency = request.form.get('currency', 'USD').strip()
        company = request.form.get('company', 'klista').strip()
        account_id = request.form.get('account_id', type=int)

        if not customer_id:
            flash('Please select a customer.', 'danger')
            return render_template('pi_edit.html', pi=pi, customers=customers,
                                   products=products, existing_items=existing_items,
                                   preload_products=preload)

        if not salesperson:
            flash('Please select a salesperson.', 'danger')
            return render_template('pi_edit.html', pi=pi, customers=customers,
                                   products=products, existing_items=existing_items,
                                   preload_products=preload)

        shipping_address = request.form.get('shipping_address', '').strip()
        try:
            shipping_cost = float(request.form.get('shipping_cost', '0').strip() or '0')
        except ValueError:
            shipping_cost = 0.0

        try:
            issue_date = datetime.strptime(issue_date_str, '%Y-%m-%d').date() if issue_date_str else pi.issue_date
        except ValueError:
            issue_date = pi.issue_date

        # Collect selected products
        selected_items = []
        total_amount = 0.0
        for product in products:
            qty_key = f'qty_{product.id}'
            selected_key = f'selected_{product.id}'
            if request.form.get(selected_key) == 'on':
                try:
                    qty = int(request.form.get(qty_key, '1').strip() or '1')
                except ValueError:
                    qty = 1
                unit_price_str = request.form.get(f'unit_price_{product.id}', '').strip()
                try:
                    unit_price = float(unit_price_str) if unit_price_str else product.unit_price
                except ValueError:
                    unit_price = product.unit_price
                if qty > 0:
                    amount = round(unit_price * qty, 2)
                    selected_items.append({
                        'product': product,
                        'quantity': qty,
                        'unit_price': unit_price,
                        'amount': amount,
                    })
                    total_amount += amount

        if not selected_items:
            flash('Please select at least one product.', 'danger')
            return render_template('pi_edit.html', pi=pi, customers=customers,
                                   products=products, existing_items=existing_items,
                                   preload_products=preload)

        total_amount = round(total_amount, 2)

        # Update PI record
        pi.customer_id = customer_id
        pi.salesperson = salesperson
        pi.currency = currency
        pi.company = company
        pi.issue_date = issue_date
        pi.payment_terms = payment_terms or '100% TT before shipment'
        bank_info_from_account = ''
        if account_id:
            acct = Account.query.get(account_id)
            if acct: bank_info_from_account = acct.bank_info()
        pi.bank_info = bank_info or bank_info_from_account
        total_amount = round(total_amount, 2)
        pi.total_amount = total_amount
        pi.shipping_cost = shipping_cost
        pi.shipping_note = request.form.get('shipping_note', '').strip()
        pi.shipping_address = shipping_address
        pi.notes = notes

        # Replace items
        PIItem.query.filter_by(pi_id=pi.id).delete()
        for item in selected_items:
            pi_item = PIItem(
                pi_id=pi.id,
                product_id=item['product'].id,
                quantity=item['quantity'],
                unit_price=item['unit_price'],
                amount=item['amount'],
            )
            db.session.add(pi_item)

        db.session.flush()

        # Regenerate PDF
        try:
            pi_full = PI.query.options(
                joinedload(PI.items).joinedload(PIItem.product)
            ).get(pi.id)
            sp = Salesperson.query.filter_by(name=pi.salesperson).first()
            sp_info = {'phone': sp.phone, 'email': sp.email} if sp else {}
            pdf_filename = generate_pi_pdf(pi_full, app.config['PDF_DIR'], sp_info)
            excel_filename = generate_pi_excel(pi_full, app.config['PDF_DIR'])
            pi.pdf_path = pdf_filename
            pi.excel_path = excel_filename
        except Exception as e:
            db.session.rollback()
            flash(f'Error generating PDF: {str(e)}', 'danger')
            return render_template('pi_edit.html', pi=pi, customers=customers,
                                   products=products, existing_items=existing_items,
                                   preload_products=preload)

        db.session.commit()
        flash(f'PI {pi.pi_number} updated and PDF regenerated.', 'success')
        return redirect(url_for('pi_list'))

    # GET request
    return render_template('pi_edit.html', pi=pi, customers=customers,
                           products=products, existing_items=existing_items,
                           preload_products=preload)


def _apply_form_to_pi(form, pi):
    """Apply form data to PI object in memory (no commit)."""
    pi.pi_number = form.get('pi_number', pi.pi_number).strip() or pi.pi_number
    pi.issue_date = datetime.strptime(form.get('issue_date', ''), '%Y-%m-%d').date() if form.get('issue_date') else pi.issue_date
    pi.salesperson = form.get('salesperson', '').strip()
    pi.currency = form.get('currency', 'USD').strip()
    pi.payment_terms = form.get('payment_terms', '').strip()
    pi.bank_info = form.get('bank_info', '').strip()
    pi.shipping_address = form.get('shipping_address', '').strip()
    pi.shipping_note = form.get('shipping_note', '').strip()
    pi.notes = form.get('notes', '').strip()
    try:
        pi.shipping_cost = float(form.get('shipping_cost', '0').strip() or '0')
    except ValueError:
        pi.shipping_cost = 0.0

    # Customer fields (in memory)
    if pi.customer:
        pi.customer.name = form.get('cust_name', '').strip() or pi.customer.name
        pi.customer.contact_person = form.get('cust_contact', '').strip()
        pi.customer.country = form.get('cust_country', '').strip()
        pi.customer.email = form.get('cust_email', '').strip()
        pi.customer.phone = form.get('cust_phone', '').strip()

    # Product items
    total = 0.0
    for item in pi.items:
        qty = int(form.get(f'qty_{item.id}', str(item.quantity)))
        price = float(form.get(f'price_{item.id}', str(item.unit_price)))
        item.quantity = qty
        item.unit_price = price
        item.amount = round(price * qty, 2)
        total += item.amount
        if item.product:
            prod_name = form.get(f'prod_name_{item.id}', '').strip()
            if prod_name:
                item.product.name = prod_name
            item.product.product_code = form.get(f'prod_code_{item.id}', '').strip()
            item.product.specification = form.get(f'prod_spec_{item.id}', '').strip()
    pi.total_amount = round(total, 2)

    # Company override — store on pi object for PDF generator to pick up
    company_name = form.get('company_name', '').strip()
    if company_name:
        pi._company_name_override = company_name
    company_addr = form.get('company_addr', '').strip()
    if company_addr:
        pi._company_addr_override = company_addr
    company_short = form.get('company_short', '').strip()
    if company_short:
        pi._company_short_override = company_short


@app.route('/pi/<int:id>/live-edit', methods=['GET', 'POST'])
@login_required
def pi_live_edit(id):
    """Online PI editor — all fields editable. Export PDF without save, or save+export."""
    pi = PI.query.options(
        joinedload(PI.customer),
        joinedload(PI.items).joinedload(PIItem.product),
    ).get_or_404(id)

    if request.method == 'POST':
        action = request.form.get('action', 'save')

        if action == 'export_pdf':
            # Generate PDF from form values — temporarily save to PI in memory, generate PDF, then rollback
            import shutil
            _apply_form_to_pi(request.form, pi)
            try:
                sp = Salesperson.query.filter_by(name=pi.salesperson).first()
                sp_info = {'phone': sp.phone, 'email': sp.email} if sp else {}
                pdf_filename = generate_pi_pdf(pi, app.config['PDF_DIR'], sp_info)
                pdf_path = os.path.join(app.config['PDF_DIR'], pdf_filename)
                # Copy to a temp location before rollback
                import tempfile
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
                shutil.copy2(pdf_path, tmp.name)
                tmp_path = tmp.name
            except Exception as e:
                db.session.rollback()
                return jsonify({'error': str(e)}), 500
            finally:
                db.session.rollback()  # Undo all form changes
                # Reload PI from DB to restore original state
                db.session.refresh(pi)
                if pi.customer:
                    db.session.refresh(pi.customer)
                for item in pi.items:
                    db.session.refresh(item)
                    if item.product:
                        db.session.refresh(item.product)

            return send_file(
                tmp_path,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=f'{pi.pi_number}_edited.pdf'
            )

        # action == 'save': original save behavior
        # Update PI header fields
        pi.issue_date = datetime.strptime(request.form.get('issue_date', ''), '%Y-%m-%d').date() if request.form.get('issue_date') else pi.issue_date
        pi.salesperson = request.form.get('salesperson', '').strip()
        pi.currency = request.form.get('currency', 'USD').strip()
        pi.payment_terms = request.form.get('payment_terms', '').strip()
        pi.bank_info = request.form.get('bank_info', '').strip()
        pi.shipping_address = request.form.get('shipping_address', '').strip()
        pi.pi_number = request.form.get('pi_number', pi.pi_number).strip() or pi.pi_number
        try:
            pi.shipping_cost = float(request.form.get('shipping_cost', '0').strip() or '0')
        except ValueError:
            pi.shipping_cost = 0.0
        pi.shipping_note = request.form.get('shipping_note', '').strip()
        pi.notes = request.form.get('notes', '').strip()

        # Update customer contact info
        customer = pi.customer
        if customer:
            customer.contact_person = request.form.get('cust_contact', '').strip()
            customer.email = request.form.get('cust_email', '').strip()
            customer.phone = request.form.get('cust_phone', '').strip()

        # Update product items
        total = 0.0
        for item in pi.items:
            qty = int(request.form.get(f'qty_{item.id}', str(item.quantity)))
            price = float(request.form.get(f'price_{item.id}', str(item.unit_price)))
            if qty > 0:
                item.quantity = qty
                item.unit_price = price
                item.amount = round(price * qty, 2)
                total += item.amount
            # Update product name/code/spec
            if item.product:
                prod_name = request.form.get(f'prod_name_{item.id}', '').strip()
                prod_code = request.form.get(f'prod_code_{item.id}', '').strip()
                prod_spec = request.form.get(f'prod_spec_{item.id}', '').strip()
                if prod_name:
                    item.product.name = prod_name
                item.product.product_code = prod_code
                item.product.specification = prod_spec

        pi.total_amount = round(total, 2)

        # Regenerate PDF
        try:
            sp = Salesperson.query.filter_by(name=pi.salesperson).first()
            sp_info = {'phone': sp.phone, 'email': sp.email} if sp else {}
            pdf_filename = generate_pi_pdf(pi, app.config['PDF_DIR'], sp_info)
            excel_filename = generate_pi_excel(pi, app.config['PDF_DIR'])
            pi.pdf_path = pdf_filename
            pi.excel_path = excel_filename
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
            return redirect(url_for('pi_live_edit', id=id))

        db.session.commit()
        flash('PI updated and PDF regenerated.', 'success')
        # Return the PDF as download
        pdf_path = os.path.join(app.config['PDF_DIR'], pi.pdf_path)
        return send_file(pdf_path, mimetype='application/pdf', as_attachment=True, download_name=f'{pi.pi_number}.pdf')

    return render_template('live_edit_pi.html', pi=pi)


# ═══════════════════════════════════════════════════════════════════════
#  ROUTES — Procurement (采购)
# ═══════════════════════════════════════════════════════════════════════

@app.route('/procurement')
@login_required
def procurement_select():
    """Select a PI to start procurement."""
    if not is_admin(): return redirect(url_for('index'))
    pis = PI.query.options(db.joinedload(PI.customer), db.joinedload(PI.items)).filter(PI.received_amount > 0).order_by(PI.id.desc()).all()
    # Check which PIs have procurement confirmed
    pi_complete = {pi.id for pi in pis if pi.procurement_confirmed}
    resp = make_response(render_template('procurement_select.html', pis=pis, pi_complete=pi_complete))
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return resp


@app.route('/procurement/<int:pi_id>')
@login_required
def procurement_page(pi_id):
    """Procurement page for a specific PI."""
    if not is_admin(): return redirect(url_for('index'))
    readonly = request.args.get('readonly', '0') == '1'
    pi = PI.query.options(
        db.joinedload(PI.customer),
        db.joinedload(PI.items).joinedload(PIItem.product)
    ).get_or_404(pi_id)
    suppliers = [s.to_dict() for s in Supplier.query.order_by(Supplier.name).all()]
    existing = {}
    for p in Procurement.query.filter_by(pi_id=pi_id).all():
        existing[str(p.pi_item_id)] = p.to_dict()
    resp = make_response(render_template('procurement.html', pi=pi, suppliers=suppliers, existing=existing, readonly=readonly))
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return resp


@app.route('/api/procurement/save', methods=['POST'])
@login_required
def api_procurement_save():
    """Save procurement items for a PI."""
    if not is_admin(): return jsonify({'success': False, 'error': 'Admin only'}), 403
    import json as _json
    data = _json.loads(request.data.decode('utf-8'))
    pi_id = data.get('pi_id')
    items = data.get('items', [])
    actual_shipping = float(data.get('actual_shipping', 0) or 0)
    proc_date = data.get('procurement_date', '') or datetime.utcnow().strftime('%Y-%m-%d')

    # Save actual shipping cost to PI
    pi = PI.query.get(pi_id)
    if pi:
        pi.actual_shipping_cost = actual_shipping

    item_ids = [it['pi_item_id'] for it in items if it.get('supplier_id')]
    if item_ids:
        Procurement.query.filter(
            Procurement.pi_id == pi_id,
            Procurement.pi_item_id.in_(item_ids)
        ).delete(synchronize_session=False)

    for it in items:
        if not it.get('supplier_id') or not it.get('quantity'):
            continue
        unit_price = float(it.get('unit_price', 0) or 0)
        quantity = int(it.get('quantity', 0) or 0)
        proc = Procurement(
            pi_id=pi_id,
            pi_item_id=it['pi_item_id'],
            supplier_id=int(it['supplier_id']),
            unit_price=unit_price,
            quantity=quantity,
            total=unit_price * quantity,
            procurement_date=proc_date,
            note=it.get('note', ''),
        )
        db.session.add(proc)

    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/procurement/<int:pi_id>/confirm', methods=['POST'])
@login_required
def api_procurement_confirm(pi_id):
    """Mark a PI's procurement as confirmed/done or unconfirm."""
    if not is_admin(): return jsonify({'success': False}), 403
    pi = PI.query.get_or_404(pi_id)
    action = request.args.get('action', 'confirm')
    pi.procurement_confirmed = (action == 'confirm')
    db.session.commit()
    return jsonify({'success': True, 'confirmed': pi.procurement_confirmed})


@app.route('/api/procurement/<int:pi_id>')
@login_required
def api_procurement_get(pi_id):
    """Get all procurement items for a PI with product names and images."""
    procs = Procurement.query.filter_by(pi_id=pi_id).all()
    result = []
    for p in procs:
        d = p.to_dict()
        if p.pi_item and p.pi_item.product:
            d['pi_item_name'] = p.pi_item.product.name
            d['pi_item_cn'] = p.pi_item.product.chinese_name or ''
            d['pi_item_image'] = p.pi_item.product.image or ''
        if p.pi_item:
            d['pi_unit_price'] = p.pi_item.unit_price
            d['pi_quantity'] = p.pi_item.quantity
        result.append(d)
    return jsonify(result)


@app.route('/backup')
@login_required
def backup_page():
    """Backup management page."""
    backup_dir = os.path.join(APP_ROOT, 'backups')
    os.makedirs(backup_dir, exist_ok=True)

    # List existing backups
    backups = []
    if os.path.exists(backup_dir):
        for f in sorted(os.listdir(backup_dir), reverse=True):
            if f.endswith('.zip') and f.startswith('backup_'):
                fp = os.path.join(backup_dir, f)
                size_mb = os.path.getsize(fp) / (1024 * 1024)
                mtime = datetime.fromtimestamp(os.path.getmtime(fp))
                backups.append({
                    'filename': f,
                    'size_mb': f'{size_mb:.1f} MB',
                    'date': mtime.strftime('%Y-%m-%d %H:%M'),
                })

    return render_template('backup.html', backups=backups)


@app.route('/backup/create', methods=['POST'])
@login_required
def backup_create():
    """Create a full backup zip (database + uploads)."""
    backup_dir = os.path.join(APP_ROOT, 'backups')
    os.makedirs(backup_dir, exist_ok=True)

    timestamp = date.today().strftime('%Y%m%d')
    backup_name = f'backup_{timestamp}.zip'
    backup_path = os.path.join(backup_dir, backup_name)

    # If backup already exists today, add sequence number
    seq = 1
    while os.path.exists(backup_path):
        backup_name = f'backup_{timestamp}_{seq:02d}.zip'
        backup_path = os.path.join(backup_dir, backup_name)
        seq += 1

    try:
        with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            # Add database
            db_dir = os.environ.get('DATABASE_DIR', os.path.join(APP_ROOT, 'instance'))
            db_path = os.path.join(db_dir, 'pi_manager.db')
            if os.path.exists(db_path):
                zf.write(db_path, 'instance/pi_manager.db')

            # Add uploads directory
            uploads_dir = app.config['UPLOAD_DIR']
            if os.path.exists(uploads_dir):
                for root, dirs, files in os.walk(uploads_dir):
                    for fn in files:
                        fp = os.path.join(root, fn)
                        arcname = os.path.relpath(fp, APP_ROOT)
                        zf.write(fp, arcname)

            # Add company config info
            info = (
                f"Backup created: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                f"Project: PI Management System\n"
                f"Company: {COMPANY_CONFIG['name']}\n"
            )
            zf.writestr('backup_info.txt', info)

        size_mb = os.path.getsize(backup_path) / (1024 * 1024)
        flash(f'Backup created: {backup_name} ({size_mb:.1f} MB)', 'success')
    except Exception as e:
        flash(f'Backup failed: {str(e)}', 'danger')

    return redirect(url_for('backup_page'))


@app.route('/backup/download/<filename>')
@login_required
def backup_download(filename):
    """Download a backup zip file."""
    backup_dir = os.path.join(APP_ROOT, 'backups')
    filepath = os.path.join(backup_dir, secure_filename(filename))
    if not os.path.exists(filepath):
        flash('Backup file not found.', 'danger')
        return redirect(url_for('backup_page'))
    return send_file(filepath, as_attachment=True, download_name=filename)


@app.route('/backup/delete/<filename>', methods=['POST'])
@login_required
def backup_delete(filename):
    """Delete a backup zip file."""
    backup_dir = os.path.join(APP_ROOT, 'backups')
    filepath = os.path.join(backup_dir, secure_filename(filename))
    if os.path.exists(filepath):
        os.remove(filepath)
        flash(f'Backup {filename} deleted.', 'success')
    return redirect(url_for('backup_page'))


# ═══════════════════════════════════════════════════════════════════════
#  Context processors — inject globals into templates
# ═══════════════════════════════════════════════════════════════════════

@app.context_processor
def inject_globals():
    user = get_current_user()
    settings = _load_settings()
    return {
        'company': COMPANY_CONFIG,
        'app_root': APP_ROOT,
        'all_salespersons': Salesperson.query.order_by(Salesperson.name).all(),
        'all_accounts': Account.query.order_by(Account.name).all(),
        'current_user': user,
        'is_admin': is_admin(),
        'exchange_rate': float(settings.get('exchange_rate', '7.0') or '7.0'),
        'today_str': date.today().strftime('%Y-%m-%d'),
    }


# ═══════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    print(f"App root: {APP_ROOT}")
    print(f"Database: {os.path.join(os.environ.get('DATABASE_DIR', os.path.join(APP_ROOT, 'instance')), 'pi_manager.db')}")
    print(f"PDF directory: {app.config['PDF_DIR']}")
    print("Starting Flask dev server on http://0.0.0.0:5000 ...")
    app.run(host='0.0.0.0', port=5000, debug=True)
