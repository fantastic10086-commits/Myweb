"""
PI Excel Generator — A4 portrait format, matching PDF layout.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

CUR_SYMBOL = {'USD': '$', 'RMB': '¥'}
BRANDS = {
    'klista': 'CHANGZHOU KLISTA INTERNATIONAL TRADE CO., LTD.',
    'qisuo':  'Changzhou QISUO Welding and Cutting Equipment Co., Ltd.',
}

# A4 portrait: ~210mm wide, print scaling
# We use column widths that sum to ~190mm for printing

def _num_to_words(n):
    ones = ['','ONE','TWO','THREE','FOUR','FIVE','SIX','SEVEN','EIGHT','NINE']
    teens = ['TEN','ELEVEN','TWELVE','THIRTEEN','FOURTEEN','FIFTEEN','SIXTEEN','SEVENTEEN','EIGHTEEN','NINETEEN']
    tens = ['','','TWENTY','THIRTY','FORTY','FIFTY','SIXTY','SEVENTY','EIGHTY','NINETY']
    def _h(n):
        if n < 10: return ones[n]
        if n < 20: return teens[n-10]
        if n < 100: return tens[n//10]+(' '+ones[n%10] if n%10 else '')
        if n < 1000: return ones[n//100]+' HUNDRED'+(' '+_h(n%100) if n%100 else '')
        return ''
    dollars = int(n); cents = int(round((n-dollars)*100))
    if dollars == 0: words = 'ZERO'
    elif dollars < 1000: words = _h(dollars)
    else:
        words = _h(dollars//1000)+' THOUSAND'
        r = dollars % 1000
        if r: words += ' '+_h(r)
    result = f'US DOLLARS {words} ONLY'
    if cents > 0: result += f' AND CENTS {cents}'
    return result


def generate_pi_excel(pi, output_dir):
    cur = getattr(pi, 'currency', 'USD') or 'USD'
    sym = CUR_SYMBOL.get(cur, '$')
    cur_label = cur
    brand = getattr(pi, 'company', 'klista') or 'klista'
    company_name = BRANDS.get(brand, BRANDS['klista'])
    filename = f"{pi.pi_number}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = pi.pi_number

    # ── Page setup for A4 portrait printing ──
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'portrait'
    from openpyxl.worksheet.page import PageMargins
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5, header=0.3, footer=0.3)
    ws.sheet_properties.pageSetUpPr = None  # reset

    # ── Styles ──
    hdr_fill = PatternFill(start_color='1a3a5c', end_color='1a3a5c', fill_type='solid')
    hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
    title_font = Font(name='Arial', bold=True, size=18, color='1a3a5c')
    bold = Font(name='Arial', bold=True, size=10)
    normal = Font(name='Arial', size=10)
    small = Font(name='Arial', size=9)
    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    ctr = Alignment(horizontal='center', vertical='center')
    rgt = Alignment(horizontal='right', vertical='center')
    lft = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Column widths for A4: A=50 B=50 (left block), C=50 (right block)
    # Actually use full width: A=6 B=40 C=15 D=20 E=20 F=20
    ws.column_dimensions['A'].width = 5    # No.
    ws.column_dimensions['B'].width = 55   # Description
    ws.column_dimensions['C'].width = 14   # QTY
    ws.column_dimensions['D'].width = 18   # Unit Price
    ws.column_dimensions['E'].width = 18   # Amount

    r = 1

    # ── Company header ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1, value=company_name)
    c.font = Font(name='Arial', bold=True, size=13, color='1a3a5c')
    c.alignment = Alignment(horizontal='center')
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1, value='No. 158 Jinchuang Road, Yaoguan Town, Changzhou City, China')
    c.font = Font(name='Arial', size=7, color='888888')
    c.alignment = Alignment(horizontal='center')
    r += 2

    # ── Title ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1, value='PROFORMA INVOICE')
    c.font = title_font; c.alignment = Alignment(horizontal='center')
    r += 2

    # ── PI Info + Customer Info (side by side) ──
    # Left: PI info
    pi_start = r
    pi_data = [
        ('PI Number:', pi.pi_number),
        ('Date:', pi.issue_date.strftime('%Y-%m-%d') if pi.issue_date else ''),
    ]
    if pi.salesperson:
        pi_data.append(('Salesperson:', pi.salesperson))

    for lb, vl in pi_data:
        ws.cell(row=r, column=1, value=lb).font = bold
        ws.cell(row=r, column=2, value=vl).font = normal
        r += 1
    r = pi_start  # reset for right column

    # Right: Customer info
    customer = pi.customer
    cust_data = [('To / Buyer:', customer.name if customer else '')]
    if customer:
        if customer.contact_person: cust_data.append(('Contact:', customer.contact_person))
        if customer.country: cust_data.append(('Country:', customer.country))
        if customer.address: cust_data.append(('Address:', customer.address))
    for lb, vl in cust_data:
        ws.cell(row=r, column=3, value=lb).font = bold
        ws.cell(row=r, column=4, value=vl).font = normal
        r += 1

    r = max(r, pi_start + len(pi_data)) + 1
    r += 1

    # ── Product Table ──
    headers = ['No.', 'Description / Item', f'QTY(pcs)', f'Unit Price({cur_label})', f'Amount({cur_label})']
    for i, h in enumerate(headers):
        c = ws.cell(row=r, column=i+1, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = ctr; c.border = thin
    r += 1

    for idx, item in enumerate(pi.items, 1):
        desc = item.product.name if item.product else ''
        if item.product and item.product.specification:
            desc += f'\n{item.product.specification}'
        vals = [idx, desc, item.quantity, item.unit_price, item.amount]
        for j, v in enumerate(vals):
            c = ws.cell(row=r, column=j+1, value=v)
            c.font = normal; c.border = thin
            if j >= 2: c.alignment = rgt; c.number_format = '#,##0.00' if j > 2 else '#,##0'
            elif j == 0: c.alignment = ctr
            else: c.alignment = lft
        r += 1
    r += 1

    # ── Total ──
    total = pi.total_amount
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(row=r, column=1, value='TOTAL:')
    c.font = Font(name='Arial', bold=True, size=12); c.alignment = rgt
    c = ws.cell(row=r, column=5, value=total)
    c.font = Font(name='Arial', bold=True, size=12); c.alignment = rgt
    c.number_format = f'{sym}#,##0.00'
    r += 1
    if cur == 'USD':
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        c = ws.cell(row=r, column=1, value=_num_to_words(total))
        c.font = Font(name='Arial', italic=True, size=8, color='666666'); c.alignment = rgt
        r += 1
    r += 1

    # ── Payment & Bank ──
    ws.cell(row=r, column=1, value='Payment Terms:').font = bold
    ws.cell(row=r, column=2, value=pi.payment_terms or '100% TT before shipment').font = normal
    r += 1
    if pi.bank_info:
        ws.cell(row=r, column=1, value='Bank Info:').font = bold
        ws.cell(row=r, column=2, value=pi.bank_info).font = normal
        ws.cell(row=r, column=2).alignment = lft
        r += 1
    ws.cell(row=r, column=1, value='Country of Origin:').font = bold
    ws.cell(row=r, column=2, value='China').font = normal
    r += 2

    # ── Signature block ──
    ws.cell(row=r, column=1, value=company_name).font = bold
    ws.cell(row=r, column=4, value='BUYER:').font = bold
    r += 3
    ws.cell(row=r, column=1, value='_________________________')
    ws.cell(row=r, column=4, value='_________________________')
    r += 1
    ws.cell(row=r, column=1, value=f'Date: {__import__("datetime").datetime.now().strftime("%Y-%m-%d")}')

    # ── Print settings ──
    ws.print_area = f'A1:E{r}'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = None

    wb.save(filepath)
    return filename
